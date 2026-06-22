# %%
import tkinter as tk
from tkinter import ttk, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from pathlib import Path
import pandas as pd
import numpy as np
import os
import sys
import csv
from datetime import datetime
import json
import re
from typing import Optional, Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
# ===== 定数 =====
DEFAULT_MIXING_KEY = "UKEIRE"
DEFAULT_HEIGHT_CAP = 2450
BASE_ONE_TIME = 187.64
MIDDLE_WORK = 3.247
BASE_PER_PAL = 52
PROC_RULE_FILENAME = "工程割当ルール.csv"
CONFIG_FILENAME = "自動仕分け設定.json"

def get_config_path() -> Path:
    """設定ファイルのパスを取得（exeと同じ場所、または.pyと同じ場所）"""
    if getattr(sys, 'frozen', False):
        # exe化された場合
        return Path(sys.executable).parent / CONFIG_FILENAME
    else:
        # 通常のPython実行
        return Path(__file__).parent / CONFIG_FILENAME

def load_config() -> dict:
    """設定ファイルを読み込む"""
    config_path = get_config_path()
    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_config(config: dict):
    """設定ファイルを保存"""
    config_path = get_config_path()
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def select_data_folder() -> Optional[Path]:
    """フォルダ選択ダイアログを表示"""
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    folder = filedialog.askdirectory(
        title="データフォルダを選択してください（出荷情報_全便_最新版.csv と 出荷場一覧.csv があるフォルダ）"
    )
    root.destroy()
    if folder:
        return Path(folder)
    return None

def get_base_dir() -> Path:
    """データフォルダのパスを取得（設定ファイルから読み込み、なければ選択ダイアログ）"""
    config = load_config()
    base_dir_str = config.get("base_dir")
    
    if base_dir_str:
        base_dir = Path(base_dir_str)
        if base_dir.exists():
            return base_dir
    
    # 設定がないか、フォルダが存在しない場合は選択ダイアログを表示
    messagebox.showinfo("初期設定", "データフォルダを選択してください。\n（出荷情報_全便_最新版.csv と 出荷場一覧.csv があるフォルダ）")
    
    while True:
        base_dir = select_data_folder()
        if base_dir is None:
            if messagebox.askyesno("確認", "フォルダが選択されていません。終了しますか？"):
                raise SystemExit("フォルダが選択されませんでした")
            continue
        
        # 必要なファイルが存在するか確認
        s_path = base_dir / "出荷情報_全便_最新版.csv"
        p_path = base_dir / "出荷場一覧.csv"
        
        if not s_path.exists() or not p_path.exists():
            missing = []
            if not s_path.exists():
                missing.append("出荷情報_全便_最新版.csv")
            if not p_path.exists():
                missing.append("出荷場一覧.csv")
            messagebox.showerror("エラー", f"選択したフォルダに必要なファイルがありません:\n{', '.join(missing)}\n\n別のフォルダを選択してください。")
            continue
        
        # 設定を保存
        config["base_dir"] = str(base_dir)
        save_config(config)
        return base_dir
_ZEN2HAN_DIGIT_COLON = str.maketrans({
    "０": "0", "１": "1", "２": "2", "３": "3", "４": "4",
    "５": "5", "６": "6", "７": "7", "８": "8", "９": "9",
    "：": ":",
})

_NAME_ALIASES = {
    "九州": "KVC",
    "TMK": "KVC",
    "日野E/H": "日野EH",
    "日野ｅ/ｈ": "日野EH",
    "日野e/h": "日野EH",
}


def _normalize_name(text: str) -> str:
    if pd.isna(text):
        return ""
    s = str(text).strip()
    return _NAME_ALIASES.get(s, s)


def _normalize_dest_name(text: str) -> str:
    return _normalize_name(text)


def _normalize_route_name(text: str) -> str:
    return _normalize_name(text)


def _normalize_hhmm(text: str) -> str:
    """
    'HH:MM' 形式に正規化。秒は切り捨て、全角数字/コロンも半角に寄せる。
    """
    if pd.isna(text):
        return ""
    s = str(text).strip().translate(_ZEN2HAN_DIGIT_COLON)
    if not s:
        return ""
    m = re.search(r"(\d{1,2}):(\d{1,2})", s)
    if not m:
        return ""
    hh, mm = m.group(1), m.group(2)
    return f"{int(hh):02d}:{int(mm):02d}"
# ===== CSV整形ユーティリティ =====
def _ensure_columns(df: pd.DataFrame, cols_order: list) -> pd.DataFrame:
    df2 = df.copy()
    for c in cols_order:
        if c not in df2.columns:
            df2[c] = ""
    return df2.reindex(columns=cols_order + [c for c in df2.columns if c not in cols_order])
def _protect_excel_injection(df: pd.DataFrame, text_cols: list) -> pd.DataFrame:
    # Excelで = + - @ 先頭の文字列が数式化されるのを防止
    def safe(s):
        if pd.isna(s):
            return s
        s = str(s)
        return "'" + s if s[:1] in ("=","+","-","@") else s
    df2 = df.copy()
    for c in text_cols:
        if c in df2.columns:
            df2[c] = df2[c].map(safe)
    return df2
# ===== 追加: A/B/C… 付与ヘルパー =====
def index_to_letters(n: int) -> str:
    """
    1→A, 26→Z, 27→AA... のExcel列名風変換
    """
    if n is None or n <= 0:
        return ""
    letters = []
    while n > 0:
        n -= 1
        letters.append(chr(65 + (n % 26)))
        n //= 26
    return "".join(reversed(letters))
def add_group_label_by_koutei_yama(
    details_df: pd.DataFrame,
    label_col: str = "GroupLabel",
    include_unset: bool = False  # True: 「未設定」工程もラベル付与対象に含める
) -> pd.DataFrame:
    """
    工程ごとに、山通番のユニーク昇順で A/B/C... を割り当てる新列を付与。
    - 同じ工程内で同じ山通番の行は同じラベルになる
    - 山通番は数値化して比較（文字列混在でも安全）
    - 既定では「未設定」工程はラベルなし。include_unset=True で付与可能
    """
    df = details_df.copy()
    if df is None or df.empty:
        df[label_col] = ""
        return df
    if "工程" not in df.columns or "山通番" not in df.columns:
        df[label_col] = ""
        return df
    # 山通番は数値化
    df["山通番"] = pd.to_numeric(df["山通番"], errors="coerce")
    # 工程キー（文字列）を作成
    df["_工程_key"] = df["工程"].astype(str).str.strip()
    # ラベル付与対象の工程セット
    if include_unset:
        target_keys = None  # すべて対象
    else:
        target_keys = {"1", "2", "3"}
    # 工程ごとにユニークな山通番一覧を作り、順位→A/B/C…に変換
    for k, sub in df.groupby("_工程_key", sort=False):
        if (target_keys is not None) and (k not in target_keys):
            continue
        uniques = sorted(sub["山通番"].dropna().unique().tolist())
        idxmap = {y: index_to_letters(i + 1) for i, y in enumerate(uniques)}
        df.loc[df["_工程_key"] == k, label_col] = df.loc[df["_工程_key"] == k, "山通番"].map(idxmap)
    df.drop(columns=["_工程_key"], inplace=True)
    return df
# ===== XLSX出力（CSVの代替） =====
def _add_table_exact(file_path, table_name):
    wb = load_workbook(file_path)
    ws = wb.active
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 1 or max_col < 1:
        wb.save(file_path); return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tbl = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tbl.tableStyleInfo = style
    names = [t.displayName for t in getattr(ws, "_tables", [])]
    if table_name in names:
        tbl.displayName = f"{table_name}_{len(names)+1}"
    ws.add_table(tbl)
    wb.save(file_path)
def export_kanban_xlsx(
    summary_df: pd.DataFrame,
    details_df: pd.DataFrame,
    out_dir: str,
    base_name: str = "工程別かんばん",
    protect_excel: bool = True,
    write_summary: bool = False  # 変更: 既定でサマリを出さない（今回は呼び出し自体しません）
) -> dict:
    """
    工程別かんばんサマリ・明細を XLSX で安全に出力し、テーブル化も行う。
    Power Automateで安全に扱えるよう、数値列は型変換＆欠損埋めを行う。
    """
    # 出力フォルダを作成
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    # 出力ファイルパス
    sum_path = os.path.join(out_dir, f"{base_name}_山サマリ.xlsx")
    det_path = os.path.join(out_dir, f"{base_name}_山明細.xlsx")
    # 想定列順（GroupLabel を追加）
    summary_cols = ["山通番", "山工程", "高さ合計", "残り容量", "キー", "納入先数", "ストア数", "RunId"]
    details_cols = ["山通番", "納入先", "工程", "工程内No", "ストア", "NONYUHIBIN", "UKEIRE",
                    "移動工数", "高さ", "サイズ種類", "RunId",
                    "SSYUKKA", "SYUKKASAKI", "SYUKKAKOKU", "HINBAN", "NONYUNO", "BANKAIHI", "BANKAIBIN",
                    "SMAKER", "SKOKU", "SEBANGO", "PLANKANBANSU", "SAKUSEITIME", "引取済",
                    "登録日時", "登録者", "納入先コード", "ローカルグループ番号", "グループ番号", "グルーピング番号",
                    "GroupLabel"]
    # Excel数式化防止対象列（GroupLabel を追加）
    text_cols_summary = ["キー", "RunId"]
    text_cols_details = ["納入先", "ストア", "NONYUHIBIN", "UKEIRE", "サイズ種類", "RunId", "GroupLabel"]
    # まずラベル列を付与（工程ごとに山通番の昇順で A/B/C…）
    details_df = add_group_label_by_koutei_yama(details_df, label_col="GroupLabel", include_unset=False)
    # 列を揃える（不足列は追加）
    s_out = _ensure_columns(summary_df, summary_cols)
    d_out = _ensure_columns(details_df, details_cols)
    # ===== 数値列の型変換＆欠損埋め =====
    # サマリ側
    for col in ["山通番", "高さ合計", "残り容量", "納入先数", "ストア数"]:
        if col in s_out.columns:
            s_out[col] = pd.to_numeric(s_out[col], errors="coerce").fillna(0).astype(int)
    # 明細側
    if "山通番" in d_out.columns:
        d_out["山通番"] = pd.to_numeric(d_out["山通番"], errors="coerce").fillna(0).astype(int)
    if "工程内No" in d_out.columns:
        d_out["工程内No"] = pd.to_numeric(d_out["工程内No"], errors="coerce").fillna(0).astype(int)
    if "移動工数" in d_out.columns:
        d_out["移動工数"] = pd.to_numeric(d_out["移動工数"], errors="coerce").fillna(0)
    if "高さ" in d_out.columns:
        d_out["高さ"] = pd.to_numeric(d_out["高さ"], errors="coerce").fillna(0)
    # Excel数式化防止
    if protect_excel:
        s_out = _protect_excel_injection(s_out, text_cols_summary)
        d_out = _protect_excel_injection(d_out, text_cols_details)
    # ===== Excel出力 =====
    if write_summary:
        s_out.to_excel(sum_path, index=False, engine="openpyxl")
        _add_table_exact(sum_path, "SummaryTable")
    d_out.to_excel(det_path, index=False, engine="openpyxl")
    _add_table_exact(det_path, "KanbanTable")
    return {"summary": sum_path if write_summary else None, "details": det_path}
# ===== サマリ生成（FutureWarning解消版） =====
def build_simple_summary(details_df: pd.DataFrame, cap: float = 2450):
    """
    明細から簡易サマリを生成（山通番ごとに 高さ合計/残り容量/納入先数/キー を集約）。
    山工程は details_df に工程列がある場合は first を採用、無ければ未設定にします。
    """
    if details_df is None or details_df.empty:
        rid = datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")
        return pd.DataFrame(columns=["山通番","山工程","高さ合計","残り容量","キー","納入先数","ストア数","RunId"]), rid
    tmp = details_df.copy()
    tmp["高さ"] = pd.to_numeric(tmp.get("高さ", 0), errors="coerce").fillna(0)
    # 高さ合計と残り容量
    h = (
        tmp.groupby("山通番", as_index=False)["高さ"]
        .sum()
        .rename(columns={"高さ": "高さ合計"})
    )
    h["残り容量"] = cap - h["高さ合計"]
    # 納入先数・ストア数（列が無い場合は0でフォールバック）
    if "納入先" in details_df.columns:
        dest_cnt = (
            details_df.groupby("山通番")["納入先"]
            .nunique()
            .reset_index(name="納入先数")
        )
    else:
        dest_cnt = details_df[["山通番"]].drop_duplicates().assign(納入先数=0)
    if "ストア" in details_df.columns:
        store_cnt = (
            details_df.groupby("山通番")["ストア"]
            .nunique()
            .reset_index(name="ストア数")
        )
    else:
        store_cnt = details_df[["山通番"]].drop_duplicates().assign(ストア数=0)
    comp = dest_cnt.merge(store_cnt, on="山通番", how="outer")
    # キー生成
    if "サイズ種類" in details_df.columns:
        stype4_only = (
            details_df.groupby("山通番")["サイズ種類"]
            .apply(lambda s: s.astype(str).eq("4").all())
            .reset_index(name="stype4_only")
        )
    else:
        stype4_only = details_df[["山通番"]].drop_duplicates().assign(stype4_only=False)
    if "NONYUHIBIN" in details_df.columns:
        nony_first = (
            details_df.groupby("山通番")["NONYUHIBIN"]
            .first()
            .reset_index(name="NONYUHIBIN_first")
        )
    else:
        nony_first = details_df[["山通番"]].drop_duplicates().assign(NONYUHIBIN_first="")
    if "UKEIRE" in details_df.columns:
        uke_first = (
            details_df.groupby("山通番")["UKEIRE"]
            .first()
            .reset_index(name="UKEIRE_first")
        )
    else:
        uke_first = details_df[["山通番"]].drop_duplicates().assign(UKEIRE_first="")
    keys = (
        stype4_only
        .merge(nony_first, on="山通番", how="left")
        .merge(uke_first, on="山通番", how="left")
    )
    keys["キー"] = np.where(
        keys["stype4_only"],
        keys["NONYUHIBIN_first"].astype(str),
        keys["UKEIRE_first"].astype(str)
    )
    keys = keys[["山通番", "キー"]]
    # 山工程
    if "工程" in details_df.columns:
        yama_proc = (
            details_df.groupby("山通番")
            .agg(山工程=("工程", "first"))
            .reset_index()
        )
    elif "山工程" in details_df.columns:
        yama_proc = (
            details_df.groupby("山通番")
            .agg(山工程=("山工程", "first"))
            .reset_index()
        )
    else:
        yama_proc = details_df[["山通番"]].drop_duplicates().assign(山工程="4")
    summary_df = (
        yama_proc.merge(h, on="山通番", how="left")
        .merge(keys, on="山通番", how="left")
        .merge(comp, on="山通番", how="left")
    )
    rid = datetime.now().strftime("%Y-%m-%dT%H:%M:%SZ")
    summary_df["RunId"] = rid
    return summary_df, rid
# ===== 混載ポリシー =====
SIZE_MIXING_POLICY = {
    "1": {"allow_mixing": True, "max_mix_groups": 3, "mixing_key": DEFAULT_MIXING_KEY},
    "4": {"allow_mixing": False},
    "default": {"allow_mixing": False}
}
def is_mixing_allowed(size_type: str) -> bool:
    st = str(size_type)
    return SIZE_MIXING_POLICY.get(st, SIZE_MIXING_POLICY["default"])["allow_mixing"]
def get_size_sort_keys(size_type: str):
    """
    種類別の並び基準（山内行の表示/処理順）
    - 種類4（武部専用）: NONYUHIBIN 昇順 → 移動工数 降順
    - 既定: 移動工数 降順 → 納入先 昇順（任意）
    """
    st = str(size_type)
    if st == "4":
        return [("NONYUHIBIN", True), ("移動工数", False)]
    else:
        return [("移動工数", False), ("納入先", True)]
def sort_dataframe_by_keys(df: pd.DataFrame, sort_keys):
    """
    sort_keys: [(col_name, ascending_bool), ...]
    存在する列のみで安全に並び替えを行う
    """
    if df is None or df.empty:
        return df
    cols = []
    asc = []
    df2 = df.copy()
    for col, up in sort_keys:
        if col in df2.columns:
            if col in ("移動工数", "高さ"):
                df2[col] = pd.to_numeric(df2[col], errors="coerce")
            cols.append(col)
            asc.append(bool(up))
    if cols:
        df2 = df2.sort_values(by=cols, ascending=asc)
    return df2
# ===== CSV読み書き =====
def read_csv_ja(path: Path) -> pd.DataFrame:
    try:
        return pd.read_csv(path, encoding="utf-8")
    except UnicodeDecodeError:
        return pd.read_csv(path, encoding="cp932")
def write_csv_ja(df: pd.DataFrame, path: Path):
    encodings = ["utf-8-sig", "cp932"]
    for enc in encodings:
        try:
            df.to_csv(path, index=False, encoding=enc)
            return
        except Exception:
            continue
    df.to_csv(path, index=False)
def load_data():
    # 設定ファイルまたはダイアログからフォルダを取得
    base_dir = get_base_dir()
    s_path = base_dir / "出荷情報_全便_最新版.csv"
    p_path = base_dir / "出荷場一覧.csv"
    if not s_path.exists() or not p_path.exists():
        raise FileNotFoundError(f"CSVが見つかりません:\n{s_path}\n{p_path}")
    return read_csv_ja(s_path), read_csv_ja(p_path)
df_shipments, df_places = load_data()
# 列名前後スペース除去
df_shipments.columns = df_shipments.columns.str.strip()
df_places.columns = df_places.columns.str.strip()
# 出荷情報側：「納入先コード」が無い場合は SYUKKASAKI をコピーして補完
if "納入先コード" not in df_shipments.columns and "SYUKKASAKI" in df_shipments.columns:
    df_shipments["納入先コード"] = df_shipments["SYUKKASAKI"].astype(str)
# 前処理（型・欠損）
for num_col in ["移動工数", "高さ", "PLANKANBANSU"]:
    if num_col in df_shipments.columns:
        if num_col == "PLANKANBANSU":
            df_shipments[num_col] = pd.to_numeric(df_shipments[num_col], errors="coerce").fillna(1).astype(int)
        else:
            df_shipments[num_col] = pd.to_numeric(df_shipments[num_col], errors="coerce").fillna(0)
for col in ["SSYUKKA","SYUKKASAKI","SYUKKAKOKU","UKEIRE","NONYUHIBIN","サイズ種類","納入先","納入先コード"]:
    if col in df_shipments.columns:
        df_shipments[col] = df_shipments[col].astype(str).fillna("")
for col in ["便名","受入","仕入先工区","納入先コード","納入先工区"]:
    if col in df_places.columns:
        df_places[col] = df_places[col].astype(str).fillna("")

# 便名の表記ゆれ補正（表示/照合を統一）
if "便名" in df_places.columns:
    df_places["便名"] = df_places["便名"].map(_normalize_route_name)
# 出荷場一覧の不足補完（必要列の確認）
required_places_cols = ["便名","受入","仕入先工区","納入先コード","納入先工区"]
missing = [c for c in required_places_cols if c not in df_places.columns]
if missing:
    if "納入先コード" in missing and "SYUKKASAKI" in df_places.columns:
        df_places["納入先コード"] = df_places["SYUKKASAKI"].astype(str)
        missing = [c for c in required_places_cols if c not in df_places.columns]
    if missing:
        raise ValueError(f"出荷場一覧.csv に必要な列が不足しています: {missing}\n必要: {required_places_cols}")
# ===== 候補抽出ユーティリティ =====
def _normalize_ukeire(val) -> str:
    """受入の正規化（数字のみなら先頭ゼロを除去して比較用に統一）"""
    if pd.isna(val):
        return ""
    s = str(val).strip()
    # 数字のみの場合は先頭ゼロを除去
    if s.isdigit():
        return str(int(s))
    return s

def _mask_for_place_row(row: pd.Series) -> pd.Series:
    place_code = row.get("納入先コード", "")
    place_ukeire = _normalize_ukeire(row["受入"])
    print(f"DEBUG: _mask_for_place_row - 仕入先工区:{row['仕入先工区']}, 納入先コード:{place_code}, 納入先工区:{row['納入先工区']}, 受入:{row['受入']} (正規化:{place_ukeire})")  # デバッグ用
    
    # 出荷情報の受入も正規化して比較
    shipment_ukeire_normalized = df_shipments["UKEIRE"].apply(_normalize_ukeire)
    
    mask = (
        (df_shipments["SSYUKKA"] == row["仕入先工区"]) &
        (df_shipments["納入先コード"] == str(place_code)) &
        (df_shipments["SYUKKAKOKU"] == row["納入先工区"]) &
        (shipment_ukeire_normalized == place_ukeire)
    )
    match_count = mask.sum()
    print(f"DEBUG: _mask_for_place_row result: {match_count} matches")  # デバッグ用
    return mask
def get_routes():
    routes = sorted(df_places["便名"].unique().tolist())
    print(f"DEBUG: get_routes() -> {routes}")  # デバッグ用
    return routes
def get_receipts_for_route(route_name):
    receipts = df_places.loc[df_places["便名"] == route_name, "受入"].unique().tolist()
    print(f"DEBUG: get_receipts_for_route('{route_name}') -> {sorted(receipts)}")  # デバッグ用
    return sorted(receipts)
def get_orders_for_route_receipt(route_name, receipt):
    """
    便名＋受入（UKEIRE）から、該当するオーダー（NONYUHIBIN）の一覧を返す。
    """
    ps = df_places[(df_places["便名"] == route_name) & (df_places["受入"] == receipt)]
    if ps.empty:
        print(f"DEBUG: get_orders_for_route_receipt('{route_name}', '{receipt}') -> no matching places")
        return []
    orders = set()
    for _, row in ps.iterrows():
        m = _mask_for_place_row(row)
        matched_orders = df_shipments.loc[m, "NONYUHIBIN"].unique().tolist()
        orders.update(matched_orders)
    result = sorted(orders, reverse=True)
    print(f"DEBUG: get_orders_for_route_receipt('{route_name}', '{receipt}') -> {len(result)} orders")
    return result
def get_orders_for_route(route_name):
    ps = df_places[df_places["便名"] == route_name]
    if ps.empty:
        return []
    mask_total = None
    for _, row in ps.iterrows():
        m = _mask_for_place_row(row)
        mask_total = m if mask_total is None else (mask_total | m)
    return sorted(df_shipments.loc[mask_total, "NONYUHIBIN"].unique().tolist(), reverse=True)
def get_receipts_for_route_order(route_name, order):
    """
    便名＋オーダー（NONYUHIBIN）から、該当する受入の一覧を返す。
    """
    print(f"DEBUG: get_receipts_for_route_order(route='{route_name}', order='{order}')")  # デバッグ用
    ps = df_places[df_places["便名"] == route_name]
    print(f"DEBUG: df_places rows for route '{route_name}': {len(ps)} rows")  # デバッグ用
    receipts = set()
    for _, row in ps.iterrows():
        print(f"DEBUG: checking place row: 受入={row.get('受入')}, 仕入先工区={row.get('仕入先工区')}, 納入先コード={row.get('納入先コード')}, 納入先工区={row.get('納入先工区')}")  # デバッグ用
        m = _mask_for_place_row(row) & (df_shipments["NONYUHIBIN"] == str(order))
        match_count = df_shipments.loc[m].shape[0]
        print(f"DEBUG: matching shipments for this place row: {match_count}")  # デバッグ用
        if match_count > 0:
            receipts.add(row["受入"])
            print(f"DEBUG: added receipt: {row['受入']}")  # デバッグ用
    result = sorted(receipts)
    print(f"DEBUG: final result: {result}")  # デバッグ用
    return result
# ===== グループ分け（上から順に積む） =====
def assign_groups_sequential(heights: pd.Series, cap: float):
    cur_g, cur_h = 1, 0.0
    out = []
    for h in heights.astype(float).to_list():
        if cur_h + h <= cap:
            out.append(cur_g); cur_h += h
        else:
            cur_g += 1; cur_h = h; out.append(cur_g)
    return out
# ===== パイプライン（集計＋種類1混載） =====
def run_pipeline(selections, height_cap, mixing_key):
    # selections: list of {"便名","受入","オーダー"}
    masks = []
    for sel in selections:
        ps = df_places[(df_places["便名"] == sel["便名"]) & (df_places["受入"] == sel["受入"])]
        if ps.empty:
            continue
        sub_mask_total = None
        for _, place_row in ps.iterrows():
            sub_mask = (
                _mask_for_place_row(place_row) &
                (df_shipments["NONYUHIBIN"] == sel["オーダー"])
            )
            sub_mask_total = sub_mask if sub_mask_total is None else (sub_mask_total | sub_mask)
        if sub_mask_total is not None:
            masks.append(sub_mask_total)
    final_mask = masks[0] if masks else pd.Series(False, index=df_shipments.index)
    for m in masks[1:]:
        final_mask |= m
    filtered = df_shipments.loc[final_mask].copy()
    # 複数枚複製（パレット単位に展開）
    if filtered.empty:
        expanded = filtered.copy()
    else:
        counts = filtered["PLANKANBANSU"].where(filtered["PLANKANBANSU"] >= 1, 1)
        idx = np.repeat(filtered.index.to_numpy(), counts.to_numpy())
        expanded = filtered.loc[idx].reset_index(drop=True)
    # 基本グループ（全サイズ種類）
    group_results, group_details = {}, {}
    if not expanded.empty:
        for size_type in expanded["サイズ種類"].astype(str).unique():
            df_sub = expanded.loc[expanded["サイズ種類"].astype(str) == str(size_type)].copy()
            sort_cols, sort_asc = (["移動工数", "NONYUHIBIN"], [False, True]) if str(size_type) == "4" else (["移動工数", "SYUKKASAKI"], [False, True])
            df_sorted = df_sub.sort_values(by=sort_cols, ascending=sort_asc).copy()
            df_sorted["グループ番号"] = assign_groups_sequential(df_sorted["高さ"], cap=height_cap)
            grp = df_sorted.groupby("グループ番号").agg(
                パレット数=("グループ番号", "count"),
                Max移動工数=("移動工数", "max"),
            ).reset_index()
            grp["引取工数"] = np.round(
                grp["Max移動工数"] + BASE_ONE_TIME + ((grp["パレット数"] - 1) * MIDDLE_WORK) + (grp["パレット数"] * BASE_PER_PAL), 0
            ).astype(int)
            group_results[str(size_type)] = grp
            group_details[str(size_type)] = df_sorted
    # 種類1の混載（納入先ごとにローカル詰め→2〜3山混載）
    size1_mixed_summary, size1_mixed_details = None, None
    if not expanded.empty and ("1" in expanded["サイズ種類"].astype(str).unique()):
        size1_df = expanded.loc[expanded["サイズ種類"].astype(str) == "1"].copy()
        if "納入先" not in size1_df.columns:
            size1_df["納入先"] = size1_df.get("納入先コード", "")
        packed_list = []
        for _, sub in size1_df.groupby("納入先", sort=False):
            sub_sorted = sub.sort_values(by=["移動工数"], ascending=[False]).copy()
            sub_sorted["ローカルグループ番号"] = assign_groups_sequential(sub_sorted["高さ"], cap=height_cap)
            packed_list.append(sub_sorted)
        size1_packed = pd.concat(packed_list, axis=0).reset_index(drop=True) if packed_list else size1_df.copy()
        # 存在するキーだけでグループ化
        group_cols = ["納入先", "ローカルグループ番号"]
        if "納入先コード" in size1_packed.columns:
            group_cols.insert(1, "納入先コード")
        aggs = {
            "高さ合計": ("高さ", "sum"),
            "Max移動工数": ("移動工数", "max"),
        }
        if mixing_key in size1_packed.columns:
            aggs[mixing_key] = (mixing_key, "first")
        group_table = size1_packed.groupby(group_cols).agg(**aggs).reset_index()
        group_table["山ID"] = np.arange(1, len(group_table) + 1)
        # 2山/3山混載（mixing_keyがある場合のみ差異条件を適用）
        used, id_map = set(), {}
        for _, g1 in group_table.sort_values("高さ合計", ascending=False).iterrows():
            id1 = int(g1["山ID"])
            if id1 in used:
                continue
            margin2 = height_cap - float(g1["高さ合計"])
            if mixing_key in group_table.columns:
                cond_mix2 = (group_table[mixing_key] != g1.get(mixing_key))
            else:
                cond_mix2 = True
            cand2 = group_table[
                (~group_table["山ID"].isin(used)) &
                (group_table["山ID"] != id1) &
                cond_mix2 &
                (group_table["高さ合計"] <= margin2)
            ].sort_values("高さ合計", ascending=False)
            if cand2.empty:
                continue
            g2 = cand2.iloc[0]
            id2 = int(g2["山ID"])
            margin3 = height_cap - float(g1["高さ合計"]) - float(g2["高さ合計"])
            if mixing_key in group_table.columns:
                cond_mix3_1 = (group_table[mixing_key] != g1.get(mixing_key))
                cond_mix3_2 = (group_table[mixing_key] != g2.get(mixing_key))
            else:
                cond_mix3_1 = True
                cond_mix3_2 = True
            cand3 = group_table[
                (~group_table["山ID"].isin(used)) &
                (~group_table["山ID"].isin([id1, id2])) &
                cond_mix3_1 &
                cond_mix3_2 &
                (group_table["高さ合計"] <= margin3)
            ].sort_values("高さ合計", ascending=False)
            used.update({id1, id2})
            id_map[id2] = id1
            if not cand3.empty:
                id3 = int(cand3.iloc[0]["山ID"])
                used.add(id3)
                id_map[id3] = id1
        # 代表山IDの割り当て
        def repr_id(x: int) -> int:
            while x in id_map:
                x = id_map[x]
            return x
        group_table["代表山ID"] = group_table["山ID"].apply(repr_id)
        rep_map = {old: i + 1 for i, old in enumerate(sorted(group_table["代表山ID"].unique()))}
        group_table["山通番"] = group_table["代表山ID"].map(rep_map).astype(int)
        # 山通番を size1_packed にマージ
        size1_with_mountain = size1_packed.merge(
            group_table[group_cols + ["山通番"]],
            on=group_cols, how="left"
        )
        # サマリ作成
        size1_mixed_summary = size1_with_mountain.groupby("山通番").agg(
            パレット数=("山通番", "count"),
            Max移動工数=("移動工数", "max")
        ).reset_index()
        size1_mixed_summary["引取工数"] = np.round(
            size1_mixed_summary["Max移動工数"] + BASE_ONE_TIME +
            ((size1_mixed_summary["パレット数"] - 1) * MIDDLE_WORK) +
            (size1_mixed_summary["パレット数"] * BASE_PER_PAL),
            0
        ).astype(int)
        # 混載キー一覧/フラグ
        if mixing_key in size1_with_mountain.columns:
            mix_list_map = {yama: "/".join(sorted(set(vals)))
                            for yama, vals in size1_with_mountain.groupby("山通番")[mixing_key]}
            size1_mixed_summary["混載キー一覧"] = size1_mixed_summary["山通番"].map(mix_list_map)
            size1_mixed_summary["混載キー種類数"] = size1_mixed_summary["混載キー一覧"].apply(lambda s: len(s.split("/")) if isinstance(s, str) and s else 0)
            size1_mixed_summary["混載フラグ"] = size1_mixed_summary["混載キー種類数"].ge(2)
        else:
            size1_mixed_summary["混載キー一覧"] = ""
            size1_mixed_summary["混載キー種類数"] = 0
            size1_mixed_summary["混載フラグ"] = False
        size1_mixed_details = size1_with_mountain.sort_values(by=["山通番", "移動工数"], ascending=[True, False])
    return filtered, expanded, group_results, group_details, size1_mixed_summary, size1_mixed_details
# ===== 追加: 全サイズの山を統合するユーティリティ =====
def build_all_mountain_details(group_details: dict, size1_mixed_details: pd.DataFrame) -> pd.DataFrame:
    """
    種類1の混載山に加え、種類1以外の基本グループも山として連番で統合する。
    - 種類1: 既存の混載結果をそのまま使用（混載あり）
    - 種類4（武部専用）: 混載なし。NONYUHIBIN を主キーに「まず同じ NONYUHIBIN を同じ山」にする。
      * ストアが違っても同じ NONYUHIBIN なら同じ山へ積む
      * 高さ上限（DEFAULT_HEIGHT_CAP）を超える場合は、同じ NONYUHIBIN で次の山を起こす
      * 山の採番は NONYUHIBIN の昇順で行う
    - その他の種類: 混載なし。既定の並び（移動工数 降順 → 納入先 昇順）で、既存グループ番号単位で山採番
    """
    frames = []
    max_id = 0
    # 種類1 混載山の取り込み（既存の山通番を尊重）
    if size1_mixed_details is not None and not size1_mixed_details.empty and ("山通番" in size1_mixed_details.columns):
        df1 = size1_mixed_details.copy()
        df1["山通番"] = pd.to_numeric(df1["山通番"], errors="coerce").fillna(0).astype(int)
        max_id = int(df1["山通番"].max()) if not df1["山通番"].empty else 0
        frames.append(df1)
    next_id = max_id + 1
    def _num(v):
        try:
            return float(v)
        except Exception:
            return 0.0
    if group_details:
        for stype, det in group_details.items():
            stype_str = str(stype)
            if stype_str == "1":
                # 種類1は既に取り込み済み（混載あり）
                continue
            if det is None or det.empty:
                continue
            det2 = det.copy()
            # 列整備（不足列は追加）
            for c in ("グループ番号", "グルーピング番号", "NONYUHIBIN", "移動工数", "高さ", "納入先",
                      "サイズ種類", "ストア", "SYUKKASAKI", "UKEIRE"):
                if c not in det2.columns:
                    if c in ("納入先", "サイズ種類", "ストア", "SYUKKASAKI", "UKEIRE", "NONYUHIBIN"):
                        det2[c] = ""
                    elif c in ("移動工数", "高さ"):
                        det2[c] = np.nan
                    else:
                        det2[c] = ""
            # 数値化
            det2["移動工数"] = pd.to_numeric(det2["移動工数"], errors="coerce")
            det2["高さ"] = pd.to_numeric(det2["高さ"], errors="coerce").fillna(0.0)
            if stype_str == "4":
                # 武部専用：NONYUHIBIN を主キーに山分け（ストアは無視）
                if "NONYUHIBIN" not in det2.columns:
                    # フォールバック
                    col_g = "グループ番号" if "グループ番号" in det2.columns else ("グルーピング番号" if "グルーピング番号" in det2.columns else None)
                    if col_g is None:
                        continue
                    det2[col_g] = pd.to_numeric(det2[col_g], errors="coerce").fillna(0).astype(int)
                    rows = []
                    for gno, sub in det2.groupby(col_g, sort=True):
                        sub2 = sub.copy()
                        sub2["山通番"] = next_id
                        next_id += 1
                        rows.append(sub2)
                    if rows:
                        frames.append(pd.concat(rows, axis=0, ignore_index=True))
                else:
                    # NONYUHIBIN昇順 → 移動工数降順で詰める
                    det2["NONYUHIBIN"] = det2["NONYUHIBIN"].astype(str).str.strip()
                    det2["_NONYUHIBIN_is_blank"] = det2["NONYUHIBIN"].eq("") | det2["NONYUHIBIN"].isna()
                    det2 = det2.sort_values(
                        by=["_NONYUHIBIN_is_blank", "NONYUHIBIN", "移動工数"],
                        ascending=[True, True, False]
                    )
                    rows = []
                    cap = DEFAULT_HEIGHT_CAP
                    for nonyu_val, grp in det2.groupby("NONYUHIBIN", sort=False):
                        current_h = 0.0
                        current_yama_id = next_id
                        for _, r in grp.iterrows():
                            h = _num(r.get("高さ", 0.0))
                            if current_h + h > cap and current_h > 0:
                                # 次の山通番（同じ NONYUHIBIN のまま継続）
                                next_id += 1
                                current_yama_id = next_id
                                current_h = 0.0
                            current_h += h
                            rr = r.copy()
                            rr["山通番"] = current_yama_id
                            rows.append(rr)
                        # この NONYUHIBIN の塊が終わったら、次の NONYUHIBIN は新規の山通番から開始
                        next_id += 1
                    if rows:
                        frames.append(pd.DataFrame(rows).drop(columns=["_NONYUHIBIN_is_blank"], errors="ignore"))
            else:
                # 既定（種類4以外）：既存グループ番号単位で山採番、並びは移動工数↓ → 納入先↑
                col_g = "グループ番号" if "グループ番号" in det2.columns else ("グルーピング番号" if "グルーピング番号" in det2.columns else None)
                if col_g is None:
                    continue
                det2[col_g] = pd.to_numeric(det2[col_g], errors="coerce").fillna(0).astype(int)
                det2 = det2.sort_values(by=["移動工数", "納入先"], ascending=[False, True])
                rows = []
                for gno, sub in det2.groupby(col_g, sort=True):
                    sub2 = sub.copy()
                    sub2["山通番"] = next_id
                    next_id += 1
                    rows.append(sub2)
                if rows:
                    frames.append(pd.concat(rows, axis=0, ignore_index=True))
    if not frames:
        return pd.DataFrame()
    all_df = pd.concat(frames, axis=0, ignore_index=True)
    # 必要列の整備と正規化
    for c in ("納入先", "山通番", "移動工数", "高さ", "サイズ種類", "UKEIRE", "ストア", "NONYUHIBIN", "SYUKKASAKI", "ローカルグループ番号"):
        if c not in all_df.columns:
            if c in ("納入先", "サイズ種類", "UKEIRE", "ストア", "NONYUHIBIN", "SYUKKASAKI"):
                all_df[c] = ""
            elif c in ("移動工数", "高さ"):
                all_df[c] = np.nan
            else:
                all_df[c] = ""
    # 文字列正規化
    all_df["納入先"] = all_df["納入先"].astype(str).str.strip()
    all_df["サイズ種類"] = all_df["サイズ種類"].astype(str).str.strip()
    # 表示順：山通番 → 移動工数↓（山内）
    if {"山通番", "移動工数"}.issubset(all_df.columns):
        all_df["移動工数"] = pd.to_numeric(all_df["移動工数"], errors="coerce")
        all_df = all_df.sort_values(["山通番", "移動工数"], ascending=[True, False]).reset_index(drop=True)
    return all_df
# ===== 表示用ユーティリティ =====
def get_dest_list_for_group(det_g: pd.DataFrame) -> list:
    # 納入先が無ければコード→SYUKKASAKIの順で代替
    if "納入先" in det_g.columns and det_g["納入先"].notna().any():
        col = "納入先"
    elif "納入先コード" in det_g.columns and det_g["納入先コード"].notna().any():
        col = "納入先コード"
    elif "SYUKKASAKI" in det_g.columns and det_g["SYUKKASAKI"].notna().any():
        col = "SYUKKASAKI"
    else:
        return []
    return sorted(set(map(str, det_g[col].tolist())))
def compute_basic_groups(group_details: dict, group_results: dict, height_cap: int) -> pd.DataFrame:
    rows = []
    if not group_details or not group_results:
        return pd.DataFrame(columns=["サイズ種類","グループ番号","パレット数","Max移動工数","引取工数","高さ合計","納入先一覧","混載"])
    for stype, det in group_details.items():
        res = group_results.get(stype)
        if det is None or det.empty or res is None or res.empty:
            continue
        for _, g in res.iterrows():
            gno = g["グループ番号"]
            det_g = det.loc[det["グループ番号"] == gno]
            hsum = float(det_g["高さ"].astype(float).sum()) if "高さ" in det_g.columns else 0.0
            dests = get_dest_list_for_group(det_g)
            rows.append({
                "サイズ種類": str(stype),
                "グループ番号": int(gno),
                "パレット数": int(g["パレット数"]),
                "Max移動工数": float(g["Max移動工数"]),
                "引取工数": int(g["引取工数"]),
                "高さ合計": int(round(hsum)),
                "納入先一覧": "/".join(dests),
                "混載": "★" if len(dests) >= 2 else ""
            })
    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=["サイズ種類","グループ番号","パレット数","Max移動工数","引取工数","高さ合計","納入先一覧","混載"])
    return df.sort_values(by=["サイズ種類","グループ番号"]).reset_index(drop=True)
# ===== 種類1混載の集約（total工数を追加） =====
def compute_mixed_groups(size1_mixed_summary: pd.DataFrame, size1_mixed_details: pd.DataFrame, height_cap: int) -> pd.DataFrame:
    if size1_mixed_summary is None or size1_mixed_details is None or size1_mixed_summary.empty or size1_mixed_details.empty:
        return pd.DataFrame(columns=["山通番","パレット数","Max移動工数","引取工数","total工数","高さ合計","混載キー種類数","混載フラグ","混載キー一覧"])
    hsum_map = (size1_mixed_details.assign(_h=size1_mixed_details["高さ"].astype(float))
                .groupby("山通番")["_h"].sum().to_dict())
    rows = []
    for _, s in size1_mixed_summary.sort_values("山通番").iterrows():
        yama = int(s["山通番"])
        hsum = float(hsum_map.get(yama, 0.0))
        max_cost = float(s["Max移動工数"])
        pick_cost = float(s["引取工数"])
        total_cost = max_cost + pick_cost
        rows.append({
            "山通番": yama,
            "パレット数": int(s["パレット数"]),
            "Max移動工数": max_cost,
            "引取工数": int(pick_cost),
            "total工数": round(total_cost, 3),
            "高さ合計": int(round(hsum)),
            "混載キー種類数": int(s.get("混載キー種類数", 0)),
            "混載フラグ": bool(s.get("混載フラグ", False)),
            "混載キー一覧": s.get("混載キー一覧", "")
        })
    return pd.DataFrame(rows)
def compute_dest_by_mountain(size1_mixed_details: pd.DataFrame, size1_mixed_summary: pd.DataFrame, height_cap: int) -> pd.DataFrame:
    if size1_mixed_summary is None or size1_mixed_details is None or size1_mixed_summary.empty or size1_mixed_details.empty:
        return pd.DataFrame(columns=["山通番","納入先数","納入先一覧","パレット数","高さ合計"])
    hsum_map = (size1_mixed_details.assign(_h=size1_mixed_details["高さ"].astype(float))
                .groupby("山通番")["_h"].sum().to_dict())
    dest_map = (size1_mixed_details.groupby("山通番")["納入先"]
                .apply(lambda s: sorted(set(map(str, s)))).to_dict())
    rows = []
    for _, s in size1_mixed_summary.iterrows():
        yama = int(s["山通番"])
        dests = dest_map.get(yama, [])
        hsum = float(hsum_map.get(yama, 0.0))
        rows.append({
            "山通番": yama,
            "納入先数": len(dests),
            "納入先一覧": "/".join(dests),
            "パレット数": int(s["パレット数"]),
            "高さ合計": int(round(hsum)),
        })
    return pd.DataFrame(rows).sort_values("山通番").reset_index(drop=True)
# 置換版 compute_proc_details（工程付与）: 納入先ベース・全サイズ山対応（並び安定化）
def compute_proc_details(mountain_details: pd.DataFrame, process_map: dict) -> pd.DataFrame:
    """
    各行の工程は納入先ベースで付与します（サイズ種類は使用しません）。
    対象: 全サイズ種類の“山”（種類1混載山＋他サイズの基本グループ山）
    """
    if mountain_details is None or mountain_details.empty:
        return pd.DataFrame()
    df = mountain_details.copy()
    # 納入先の正規化（なければコード/SYUKKASAKIで代替）
    if "納入先" not in df.columns or not df["納入先"].notna().any():
        df["納入先"] = df.get("納入先コード", df.get("SYUKKASAKI", "")).astype(str)
    df["納入先"] = df["納入先"].astype(str).fillna("").map(_normalize_dest_name)
    def to_label(v):
        return str(int(v)) if isinstance(v, (int, np.integer)) and v in (1, 2, 3) else "4"
    # 納入先→工程の割当
    df["工程"] = df["納入先"].map(lambda d: to_label(process_map.get(d, None)))
    # 並び：山通番↑ → 工程↑ → 移動工数↓ → 納入先↑（存在列のみ）
    sort_plan = [("山通番", True), ("工程", True), ("移動工数", False), ("納入先", True)]
    sort_cols, ascending = [], []
    for col, asc in sort_plan:
        if col in df.columns:
            sort_cols.append(col); ascending.append(asc)
    if sort_cols:
        if "移動工数" in sort_cols:
            df["移動工数"] = pd.to_numeric(df["移動工数"], errors="coerce")
        df = df.sort_values(by=sort_cols, ascending=ascending)
    # 工程内No
    grp_cols = [c for c in ["山通番", "工程"] if c in df.columns]
    if grp_cols:
        df["工程内No"] = df.groupby(grp_cols).cumcount() + 1
    else:
        df["工程内No"] = 1
    return df
# 置換版 山単位の工程決定（武部のみ強制3＋優先度/多数決）
def compute_mountain_process(proc_details: pd.DataFrame, strategy: str = "priority", force_takebe_to_3: bool = True) -> pd.DataFrame:
    """
    山工程の決定:
    - 任意: 山が“武部のみ”なら工程=3を強制（納入先ルールが未設定でも3）
    - それ以外は行工程（納入先ベース）を集計し、strategy（優先度/多数決）で決定
    """
    if proc_details is None or proc_details.empty:
        return pd.DataFrame(columns=["山通番", "山工程"])
    df = proc_details.copy()
    df["工程_norm"] = df["工程"].astype(str)
    df["工程_norm"] = df["工程_norm"].where(df["工程_norm"].isin(["1", "2", "3", "4"]), "4")
    rows = []
    for yama, sub in df.groupby("山通番"):
        # 任意：武部のみの山は強制3
        if force_takebe_to_3 and ("納入先" in sub.columns):
            dests = set(sub["納入先"].astype(str).str.strip().unique())
            if dests and dests == {"武部"}:
                rows.append({"山通番": int(yama), "山工程": "3"})
                continue
        counts = sub["工程_norm"].value_counts()
        if strategy == "priority":
            if counts.get("1", 0) > 0:
                y_proc = "1"
            elif counts.get("2", 0) > 0:
                y_proc = "2"
            elif counts.get("3", 0) > 0:
                y_proc = "3"
            else:
                y_proc = "4"
        else:  # majority
            best = None; best_n = -1
            for p in ["1", "2", "3"]:
                n = counts.get(p, 0)
                if n > best_n:
                    best, best_n = p, n
            y_proc = best if best_n > 0 else "4"
        rows.append({"山通番": int(yama), "山工程": y_proc})
    return pd.DataFrame(rows).sort_values("山通番").reset_index(drop=True)


# ===== 入車時間ベースの工程自動割り振り =====
def assign_processes_by_arrival_time(
    proc_details: pd.DataFrame,
    master_df: pd.DataFrame,
    num_processes: int = 4
) -> pd.DataFrame:
    """
    入車時間マスタを元に、入車時間に間に合うように1工程から優先的に山を割り振る。
    
    ロジック:
    1. 各山の締め切り時間（入車時間）と引取工数を算出
    2. 締め切り時間順に山をソート
    3. 各山について1工程から順に「その工程で間に合うか」をチェック
    4. 間に合う最初の工程に割り当て（間に合わなければ次の工程へ）
    5. どの工程でも間に合わなければ最後の工程（4工程）に割り当て
    
    Args:
        proc_details: 工程明細DataFrame（山通番、納入先、NONYUHIBIN等を含む）
        master_df: 入車時間マスタ（OData_納入先, NONYUHIBIN, 入車時間）
        num_processes: 使用する工程数（既定4）
    
    Returns:
        DataFrame[山通番, 山工程]
    """
    if proc_details is None or proc_details.empty:
        return pd.DataFrame(columns=["山通番", "山工程"])
    
    if master_df is None or master_df.empty:
        # マスタがない場合は従来の優先度ベースで割り振り
        return compute_mountain_process(proc_details, strategy="priority", force_takebe_to_3=False)
    
    # 休憩時間の定義（秒単位）
    BREAK_TIMES = [
        (8*3600 + 30*60, 8*3600 + 40*60),    # 8:30~8:40
        (10*3600 + 40*60, 11*3600 + 25*60),  # 10:40~11:25
        (12*3600 + 55*60, 13*3600 + 5*60),   # 12:55~13:05
        (18*3600 + 45*60, 18*3600 + 55*60),  # 18:45~18:55
        (20*3600 + 55*60, 21*3600 + 40*60),  # 20:55~21:40
        (23*3600 + 10*60, 23*3600 + 20*60),  # 23:10~23:20
    ]
    
    def _time_to_seconds(hhmm: str) -> Optional[int]:
        """HH:MM形式を秒に変換"""
        s = _normalize_hhmm(hhmm)
        if not s:
            return None
        try:
            hh, mm = s.split(":", 1)
            return int(hh) * 3600 + int(mm) * 60
        except Exception:
            return None
    
    def _calc_work_end_with_breaks(start_secs: int, work_duration_secs: int) -> int:
        """
        作業開始時間と作業時間から、休憩を考慮した作業終了時間を計算
        """
        current = start_secs
        remaining = work_duration_secs
        
        while remaining > 0:
            # 次に来る休憩時間を探す
            next_break_start = None
            next_break_end = None
            for bs, be in BREAK_TIMES:
                if current < bs:
                    # まだ始まっていない休憩
                    if next_break_start is None or bs < next_break_start:
                        next_break_start = bs
                        next_break_end = be
                elif bs <= current < be:
                    # 休憩中なら休憩終了後にスキップ
                    current = be
                    continue
            
            if next_break_start is None:
                # これ以上休憩がない
                current += remaining
                remaining = 0
            else:
                # 次の休憩までに作業完了できるか
                time_until_break = next_break_start - current
                if remaining <= time_until_break:
                    current += remaining
                    remaining = 0
                else:
                    # 休憩をまたぐ
                    remaining -= time_until_break
                    current = next_break_end
        
        return current
    
    # 入車時間マスタからマッピングを作成
    master = master_df.copy()
    master["OData_納入先"] = master["OData_納入先"].astype(str).str.strip()
    master["NONYUHIBIN"] = master["NONYUHIBIN"].astype(str).str.strip().str.translate(_ZEN2HAN_DIGIT_COLON)
    master["入車時間"] = master["入車時間"].astype(str).str.strip()
    master_map = {(r["OData_納入先"], r["NONYUHIBIN"]): r["入車時間"] for _, r in master.iterrows()}
    
    # 納入先ごとの入車時間グループ（武部の前グループ検索用）
    vendor_time_groups: Dict[str, Dict[int, List[str]]] = {}
    for (v, bin_no), pickup_time in master_map.items():
        if v not in vendor_time_groups:
            vendor_time_groups[v] = {}
        normalized = _normalize_hhmm(pickup_time)
        if normalized:
            try:
                hh, mm = normalized.split(":", 1)
                mins = int(hh) * 60 + int(mm)
                if mins not in vendor_time_groups[v]:
                    vendor_time_groups[v][mins] = []
                vendor_time_groups[v][mins].append(bin_no)
            except:
                pass
    
    vendor_sorted_groups: Dict[str, List[Tuple[int, List[str]]]] = {}
    for v, time_dict in vendor_time_groups.items():
        sorted_times = sorted(time_dict.keys())
        vendor_sorted_groups[v] = [(t, time_dict[t]) for t in sorted_times]
    
    def _get_prev_group_time(vendor: str, current_mins: int) -> Optional[int]:
        if vendor not in vendor_sorted_groups:
            return None
        groups = vendor_sorted_groups[vendor]
        prev_time = None
        for time_mins, _ in groups:
            if time_mins >= current_mins:
                break
            prev_time = time_mins
        return prev_time
    
    df = proc_details.copy()
    df["移動工数"] = pd.to_numeric(df.get("移動工数", np.nan), errors="coerce")
    
    # 各山の情報を集計
    mountain_info = []
    for yama, sub in df.groupby("山通番", sort=True):
        yama_int = int(yama)
        pal = int(sub.shape[0])
        max_cost = float(sub["移動工数"].max()) if sub["移動工数"].notna().any() else 0.0
        # 引取工数（秒単位）
        pick_cost_secs = int(np.round(
            max_cost + BASE_ONE_TIME + ((pal - 1) * MIDDLE_WORK) + (pal * BASE_PER_PAL), 0
        ))
        
        # 入車時間（締め切り）を取得 - 山内の最も早い入車時間を採用
        deadline_secs = None
        start_time_secs = None
        
        for _, row in sub.iterrows():
            vendor = _normalize_dest_name(str(row.get("納入先", row.get("OData_納入先", ""))))
            nony = str(row.get("NONYUHIBIN", "")).strip().translate(_ZEN2HAN_DIGIT_COLON)
            order2 = nony[-2:] if len(nony) >= 2 else ""
            
            if not vendor or not order2:
                continue
            
            pickup = master_map.get((vendor, order2), "")
            if not pickup:
                continue
            
            pickup_secs = _time_to_seconds(pickup)
            if pickup_secs is None:
                continue
            
            # 入車時間（締め切り）を更新
            if deadline_secs is None or pickup_secs < deadline_secs:
                deadline_secs = pickup_secs
            
            # 引取開始時間を計算（既存ロジックを踏襲）
            if vendor == "武部":
                mins = pickup_secs // 60
                prev_group_time = _get_prev_group_time(vendor, mins)
                if prev_group_time is not None:
                    st = (prev_group_time + 10) * 60  # 分→秒
                else:
                    st = pickup_secs + 10 * 60
            else:
                try:
                    current_bin = int(order2)
                    if current_bin > 1:
                        prev_bin = f"{current_bin - 1:02d}"
                        prev_pickup = master_map.get((vendor, prev_bin), "")
                        if prev_pickup:
                            prev_secs = _time_to_seconds(prev_pickup)
                            if prev_secs is not None:
                                st = prev_secs + 10 * 60
                            else:
                                st = pickup_secs + 10 * 60
                        else:
                            st = pickup_secs + 10 * 60
                    else:
                        st = pickup_secs + 10 * 60
                except (ValueError, TypeError):
                    st = pickup_secs + 10 * 60
            
            if start_time_secs is None or st < start_time_secs:
                start_time_secs = st
        
        mountain_info.append({
            "山通番": yama_int,
            "引取工数_秒": pick_cost_secs,
            "締め切り_秒": deadline_secs,
            "開始時間_秒": start_time_secs,
        })
    
    if not mountain_info:
        return pd.DataFrame(columns=["山通番", "山工程"])
    
    # 締め切り時間順にソート（締め切りがないものは後ろ）
    mountain_info.sort(key=lambda x: (x["締め切り_秒"] is None, x["締め切り_秒"] or float('inf')))
    
    # デバッグ用: 山情報を表示
    def _secs_to_hhmm(secs):
        if secs is None:
            return "None"
        hh = secs // 3600
        mm = (secs % 3600) // 60
        return f"{hh:02d}:{mm:02d}"
    
    print("\n=== 工程割り振りデバッグ情報 ===")
    for m in mountain_info:
        print(f"山{m['山通番']}: 引取工数={m['引取工数_秒']}秒, "
              f"締め切り={_secs_to_hhmm(m['締め切り_秒'])}, "
              f"開始時間={_secs_to_hhmm(m['開始時間_秒'])}")
    
    # 各工程の作業終了時間を追跡（初期値は0:00）
    proc_end_times = {str(p): 0 for p in range(1, num_processes + 1)}
    
    results = []
    for m in mountain_info:
        yama = m["山通番"]
        work_duration = m["引取工数_秒"]
        deadline = m["締め切り_秒"]
        start_time = m["開始時間_秒"]
        
        assigned_proc = str(num_processes)  # デフォルトは最後の工程
        assignment_reason = "デフォルト（4工程）"
        
        if deadline is None:
            # 締め切りがない場合は最も空いている工程に割り当て
            earliest_proc = min(proc_end_times.keys(), key=lambda p: proc_end_times[p])
            assigned_proc = earliest_proc
            assignment_reason = "締め切りなし→最も空いている工程"
        else:
            # 1工程から順にチェック
            for p in range(1, num_processes + 1):
                proc_key = str(p)
                proc_available = proc_end_times[proc_key]
                
                # この工程での作業開始時間（工程が空いてから or 山の開始時間の遅い方）
                actual_start = max(proc_available, start_time or 0)
                
                # 休憩を考慮した作業終了時間
                work_end = _calc_work_end_with_breaks(actual_start, work_duration)
                
                print(f"  山{yama} {p}工程チェック: 工程空き={_secs_to_hhmm(proc_available)}, "
                      f"実際開始={_secs_to_hhmm(actual_start)}, "
                      f"作業終了={_secs_to_hhmm(work_end)}, "
                      f"締め切り={_secs_to_hhmm(deadline)}, "
                      f"間に合う={'Yes' if work_end <= deadline else 'No'}")
                
                # 締め切りに間に合うか
                if work_end <= deadline:
                    assigned_proc = proc_key
                    proc_end_times[proc_key] = work_end
                    assignment_reason = f"{p}工程で締め切りに間に合う"
                    break
            else:
                # どの工程でも間に合わない場合、最も空いている工程に割り当て
                # （最も早く作業を開始できる工程）
                earliest_proc = min(proc_end_times.keys(), key=lambda p: proc_end_times[p])
                assigned_proc = earliest_proc
                actual_start = max(proc_end_times[assigned_proc], start_time or 0)
                work_end = _calc_work_end_with_breaks(actual_start, work_duration)
                proc_end_times[assigned_proc] = work_end
                assignment_reason = f"どの工程でも間に合わない→{earliest_proc}工程（最も空き）"
        
        print(f"  → 山{yama} 割り当て: {assigned_proc}工程 ({assignment_reason})")
        results.append({"山通番": yama, "山工程": assigned_proc})
    
    print("=== 工程割り振り完了 ===\n")
    
    return pd.DataFrame(results).sort_values("山通番").reset_index(drop=True)


# 差し替え：工程サマリ（山を1工程に振り分けたカウント）
def compute_proc_summary(proc_details: pd.DataFrame) -> pd.DataFrame:
    if proc_details is None or proc_details.empty:
        return pd.DataFrame(columns=["山通番", "工程1", "工程2", "工程3", "4工程", "合計"])
    y_df = compute_mountain_process(proc_details, strategy="priority")
    out_rows = []
    for _, r in y_df.iterrows():
        yama = int(r["山通番"])
        lab = str(r["山工程"])
        row = {"山通番": yama, "工程1": 0, "工程2": 0, "工程3": 0, "4工程": 0}
        if lab in ("1", "2", "3"):
            row[f"工程{lab}"] = 1
        else:
            row["4工程"] = 1
        row["合計"] = 1
        out_rows.append(row)
    out = pd.DataFrame(out_rows)
    return out[["山通番", "工程1", "工程2", "工程3", "4工程", "合計"]].sort_values("山通番").reset_index(drop=True)
# Total工数テーブル
def compute_total_work_table(group_details: dict,
                             group_results: dict,
                             height_cap: int,
                             size1_mixed_summary: pd.DataFrame,
                             size1_mixed_details: pd.DataFrame) -> pd.DataFrame:
    df_basic = compute_basic_groups(group_details, group_results, height_cap)
    df_mix = compute_mixed_groups(size1_mixed_summary, size1_mixed_details, height_cap)
    rows = []
    # 基本グループ
    if df_basic is not None and not df_basic.empty:
        for _, r in df_basic.iterrows():
            max_cost = float(r.get("Max移動工数", 0))
            pick_cost = float(r.get("引取工数", 0))
            total = max_cost + pick_cost
            rows.append({
                "区分": "基本グループ",
                "キー": f"{r.get('サイズ種類')}-{r.get('グループ番号')}",
                "パレット数": int(r.get("パレット数", 0)),
                "total工数": round(total, 3),
                "納入先一覧": r.get("納入先一覧", "")
            })
    # 山ごとの納入先一覧（種類1詳細から）
    dest_map = {}
    if size1_mixed_details is not None and not size1_mixed_details.empty:
        if "納入先" in size1_mixed_details.columns and "山通番" in size1_mixed_details.columns:
            dest_map = (size1_mixed_details.groupby("山通番")["納入先"]
                        .apply(lambda s: "/".join(sorted(set(map(str, s))))).to_dict())
    # 種類1混載
    if df_mix is not None and not df_mix.empty:
        for _, r in df_mix.iterrows():
            yama = int(r.get("山通番"))
            total = float(r.get("total工数", 0))
            rows.append({
                "区分": "種類1混載",
                "キー": f"山{yama}",
                "パレット数": int(r.get("パレット数", 0)),
                "total工数": round(total, 3),
                "納入先一覧": dest_map.get(yama, "")
            })
    df_total = pd.DataFrame(rows)
    cols = ["区分","キー","パレット数","total工数","納入先一覧"]
    if df_total.empty:
        return pd.DataFrame(columns=cols)
    return df_total[cols].sort_values(by=["区分", "キー"]).reset_index(drop=True)
# ===== Power Apps 互換の GroupedData JSON（工程=1/2/3コレクション向け） =====
def build_groupeddata_json_for_powerapps(det_y: pd.DataFrame) -> str:
    """
    Power Apps の JSON(ShowColumns(GroupedData, OData__x30b9__x30c8__x30a2_, NONYUHIBIN, UKEIRE,
                        OData__x7d0d__x5165__x5148_, SEBANGO, 番号, 引取済))
    と同等の JSON 配列文字列。
    """
    if det_y is None or det_y.empty:
        return "[]"
    df = det_y.copy()
    if "OData__x30b9__x30c8__x30a2_" not in df.columns:
        df["OData__x30b9__x30c8__x30a2_"] = df.get("ストア", df.get("SYUKKASAKI", "")).astype(str)
    if "OData__x7d0d__x5165__x5148_" not in df.columns:
        df["OData__x7d0d__x5165__x5148_"] = df.get("納入先", "").astype(str)
    if "引取済" not in df.columns:
        df["引取済"] = ""
    # 納入先で揃えて見やすくする（混載でも同じ納入先が連続する）
    sort_plan = [
        ("OData__x7d0d__x5165__x5148_", True),
        ("OData__x30b9__x30c8__x30a2_", True),
        ("工程内No", True),
        ("移動工数", False),
    ]
    by, asc = [], []
    for c, a in sort_plan:
        if c in df.columns:
            by.append(c); asc.append(a)
    if by:
        if "移動工数" in by:
            df["移動工数"] = pd.to_numeric(df["移動工数"], errors="coerce")
        df = df.sort_values(by=by, ascending=asc)
    df = df.reset_index(drop=True)
    df["番号"] = np.arange(1, len(df) + 1)
    cols = ["OData__x30b9__x30c8__x30a2_", "NONYUHIBIN", "UKEIRE",
            "OData__x7d0d__x5165__x5148_", "SEBANGO", "番号", "引取済"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    records = df[cols].astype(object).where(pd.notna(df[cols]), "").to_dict(orient="records")
    return json.dumps(records, ensure_ascii=False)
def make_powerapps_like_collection(
    proc_details: pd.DataFrame,
    mountain_proc_map: Dict[int, str],
    process_code: str = "1",          # "1"/"2"/"3"
    process_label: str = "1工程",
    pickup_start_times: Optional[Dict[int, str]] = None  # {山通番: "HH:MM"}
) -> pd.DataFrame:
    """
    Power Apps の ForAll / Collect 相当の 1工程用コレクションを生成（DataFrame）
    出力列: 工程, ID(山通番), GroupedData(JSON文字列), Max移動工数, グループ番号, パレット数, 引取工数, 引取開始時間
    """
    cols_out = ["工程", "ID", "GroupedData", "Max移動工数",
                "グループ番号", "パレット数", "引取工数", "引取開始時間"]
    if proc_details is None or proc_details.empty or not mountain_proc_map:
        return pd.DataFrame(columns=cols_out)
    df = proc_details.copy()
    df["移動工数"] = pd.to_numeric(df.get("移動工数", np.nan), errors="coerce")
    target_yamas = [int(y) for y, p in mountain_proc_map.items() if str(p) == str(process_code)]
    if not target_yamas:
        return pd.DataFrame(columns=cols_out)
    df_y = df[df["山通番"].isin(target_yamas) & (df["工程"].astype(str) == str(process_code))].copy()
    if df_y.empty:
        df_y = df[df["山通番"].isin(target_yamas)].copy()
    rows = []
    for yama, sub in df_y.groupby("山通番", sort=True):
        gd_json = build_groupeddata_json_for_powerapps(sub)
        pal = int(sub.shape[0])
        max_cost = float(sub["移動工数"].max()) if sub["移動工数"].notna().any() else 0.0
        pick_cost = int(np.round(
            max_cost + BASE_ONE_TIME + ((pal - 1) * MIDDLE_WORK) + (pal * BASE_PER_PAL), 0
        ))
        start_time = ""
        if isinstance(pickup_start_times, dict):
            start_time = str(pickup_start_times.get(int(yama), "")).strip()
        rows.append({
            "工程": process_label,
            "ID": int(yama),
            "GroupedData": gd_json,
            "Max移動工数": max_cost,
            "グループ番号": int(yama),
            "パレット数": pal,
            "引取工数": pick_cost,
            "引取開始時間": start_time
        })
    out = pd.DataFrame(rows, columns=cols_out)
    if out.empty:
        return out
    out["ID"] = out["ID"].astype(int)
    out["グループ番号"] = out["グループ番号"].astype(int)
    out["パレット数"] = out["パレット数"].astype(int)
    out["引取工数"] = out["引取工数"].astype(int)
    out["Max移動工数"] = out["Max移動工数"].astype(float)
    return out
# ===== 工程ごとの山番（工程内山番）採番 ＋ 1山=1行（GroupedData含む） =====
def compute_per_process_mountain_seq(mountain_proc: pd.DataFrame, include_unset: bool = True):
    """
    山工程ごとに 山通番 を昇順に並べ、工程内連番(1,2,...)を採番する。
    戻り値:
      - seq_map: {山通番(int): 工程内山番(int)}
      - df_seq:  DataFrame[山通番, 山工程, 工程内山番]
    """
    if mountain_proc is None or mountain_proc.empty:
        return {}, pd.DataFrame(columns=["山通番", "山工程", "工程内山番"])
    df = mountain_proc.copy()
    df["山工程"] = df["山工程"].astype(str)
    if not include_unset:
        df = df[df["山工程"].isin(["1","2","3"])].copy()
    out_rows = []
    for proc, sub in df.groupby("山工程", sort=True):
        sub2 = sub.sort_values("山通番").copy()
        sub2["工程内山番"] = np.arange(1, len(sub2) + 1)
        out_rows.append(sub2[["山通番", "山工程", "工程内山番"]])
    if not out_rows:
        return {}, pd.DataFrame(columns=["山通番", "山工程", "工程内山番"])
    df_seq = pd.concat(out_rows, axis=0, ignore_index=True)
    seq_map = dict(zip(df_seq["山通番"].astype(int), df_seq["工程内山番"].astype(int)))
    return seq_map, df_seq
def build_groupeddata_json_for_mountain(sub_rows: pd.DataFrame) -> str:
    """
    ある山の（対象工程で絞った）行から GroupedData(JSON配列文字列) を作る。
    出力キー:
      - OData__x30b9__x30c8__x30a2_（ストア。無ければ SYUKKASAKI）
      - NONYUHIBIN, UKEIRE
      - OData__x7d0d__x5165__x5148_（納入先）
      - SEBANGO
      - 番号（1..N: 表示順ベース）
      - 引取済（無ければ空）
    """
    if sub_rows is None or sub_rows.empty:
        return "[]"
    df = sub_rows.copy()
    if "OData__x30b9__x30c8__x30a2_" not in df.columns:
        df["OData__x30b9__x30c8__x30a2_"] = df.get("ストア", df.get("SYUKKASAKI", "")).astype(str)
    if "OData__x7d0d__x5165__x5148_" not in df.columns:
        df["OData__x7d0d__x5165__x5148_"] = df.get("納入先", "").astype(str)
    if "引取済" not in df.columns:
        df["引取済"] = ""
    # 納入先で揃えて見やすくする（混載でも同じ納入先が連続する）
    sort_plan = [
        ("OData__x7d0d__x5165__x5148_", True),
        ("OData__x30b9__x30c8__x30a2_", True),
        ("工程内No", True),
        ("移動工数", False),
    ]
    by, asc = [], []
    for c, a in sort_plan:
        if c in df.columns:
            by.append(c); asc.append(a)
    if by:
        if "移動工数" in by:
            df["移動工数"] = pd.to_numeric(df["移動工数"], errors="coerce")
        df = df.sort_values(by=by, ascending=asc)
    df = df.reset_index(drop=True)
    df["番号"] = np.arange(1, len(df) + 1)
    cols = ["OData__x30b9__x30c8__x30a2_", "NONYUHIBIN", "UKEIRE",
            "OData__x7d0d__x5165__x5148_", "SEBANGO", "番号", "引取済"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    recs = df[cols].astype(object).where(pd.notna(df[cols]), "").to_dict(orient="records")
    return json.dumps(recs, ensure_ascii=False)
def build_per_process_mountain_rows(proc_details: pd.DataFrame,
                                    mountain_proc: pd.DataFrame,
                                    include_unset: bool = True) -> pd.DataFrame:
    """
    山工程ごとに山番を振り直し（工程内山番）、
    1山=1行のレコードを構成して返す。
    """
    cols_out = ["工程", "工程内山番", "山通番", "キー",
                "パレット数", "Max移動工数", "引取工数", "GroupedData"]
    if proc_details is None or proc_details.empty or mountain_proc is None or mountain_proc.empty:
        return pd.DataFrame(columns=cols_out)
    # 工程内山番の採番
    seq_map, df_seq = compute_per_process_mountain_seq(mountain_proc, include_unset=include_unset)
    if not seq_map:
        return pd.DataFrame(columns=cols_out)
    det = proc_details.copy()
    det["移動工数"] = pd.to_numeric(det.get("移動工数", np.nan), errors="coerce")
    # 山工程を det に付与
    ymap = dict(zip(mountain_proc["山通番"].astype(int), mountain_proc["山工程"].astype(str)))
    det["山工程"] = det["山通番"].map(ymap).astype(str)
    rows = []
    for yama, y_proc in ymap.items():
        if (not include_unset) and (y_proc not in ("1","2","3")):
            continue
        sub = det[(det["山通番"] == yama) & (det["工程"].astype(str) == y_proc)].copy()
        if sub.empty:
            sub = det[det["山通番"] == yama].copy()
        pal = int(sub.shape[0])  # パレット数
        max_cost = float(sub["移動工数"].max()) if sub["移動工数"].notna().any() else 0.0
        pick_cost = int(np.round(max_cost + BASE_ONE_TIME + ((pal - 1) * MIDDLE_WORK) + (pal * BASE_PER_PAL), 0))
        gd_json = build_groupeddata_json_for_mountain(sub)
        seq = int(seq_map.get(int(yama), 0))
        proc_label = f"{y_proc}工程" if y_proc in ("1","2","3") else "4工程"
        rows.append({
            "工程": proc_label,
            "工程内山番": seq,
            "山通番": int(yama),
            "キー": f"{y_proc}-{seq}" if y_proc in ("1","2","3") else f"4-{seq}",
            "パレット数": pal,
            "Max移動工数": max_cost,
            "引取工数": pick_cost,
            "GroupedData": gd_json
        })
    out = pd.DataFrame(rows, columns=cols_out)
    if out.empty:
        return out
    out["工程内山番"] = pd.to_numeric(out["工程内山番"], errors="coerce").fillna(0).astype(int)
    out["山通番"] = pd.to_numeric(out["山通番"], errors="coerce").fillna(0).astype(int)
    out["パレット数"] = pd.to_numeric(out["パレット数"], errors="coerce").fillna(0).astype(int)
    out["引取工数"] = pd.to_numeric(out["引取工数"], errors="coerce").fillna(0).astype(int)
    out["Max移動工数"] = pd.to_numeric(out["Max移動工数"], errors="coerce").fillna(0.0).astype(float)
    # 並び：工程→工程内山番→山通番
    out = out.sort_values(by=["工程", "工程内山番", "山通番"]).reset_index(drop=True)
    out = _protect_excel_injection(out, ["工程", "キー", "GroupedData"])
    return out
def export_per_process_mountains_xlsx(perproc_df: pd.DataFrame, out_dir: str,
                                      base_name: str = "工程別_山1レコード") -> str:
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    path = os.path.join(out_dir, f"{base_name}.xlsx")
    perproc_df.to_excel(path, index=False, engine="openpyxl")
    _add_table_exact(path, "PerProcessMountains")
    return path
# ===== 入車時間マスタ＆GroupedData処理 =====
def get_master_path() -> Path:
    """入車時間マスタのパスを取得（プログラムと同階層）"""
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent / "入車時間マスタ.xlsx"
    else:
        return Path(__file__).parent / "入車時間マスタ.xlsx"

def save_pickup_time_master_xlsx(df: pd.DataFrame, master_path: Path):
    """
    入車時間マスタをExcelファイルに保存
    """
    df_save = df.copy()
    # 必須列の確認
    expected_cols = ["OData_納入先", "NONYUHIBIN", "入車時間"]
    for col in expected_cols:
        if col not in df_save.columns:
            df_save[col] = ""
    df_save = df_save[expected_cols]
    df_save.to_excel(master_path, index=False, engine="openpyxl", sheet_name="入車時間マスタ")

def load_pickup_time_master_xlsx(master_path: Path) -> pd.DataFrame:
    """
    入車時間マスタ（先頭シート）を読み込み、キー列を正規化して返す。
    - 全角数字/コロンは半角へ
    - NONYUHIBIN は2桁ゼロ埋め
    - 入車時間は 'HH:MM' に整形（秒は切り捨て）
    """
    if not master_path.exists():
        # ファイルが存在しない場合は空のDataFrameを返す
        return pd.DataFrame(columns=["OData_納入先", "NONYUHIBIN", "入車時間"])
    df = pd.read_excel(master_path, sheet_name=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    expected_cols = ["OData_納入先", "NONYUHIBIN", "入車時間"]
    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        raise ValueError(f"入車時間マスタに必要な列がありません: {', '.join(missing)}")
    df = df[expected_cols].copy()
    df["OData_納入先"] = df["OData_納入先"].astype(str).str.strip()
    nony = df["NONYUHIBIN"].astype(str).str.translate(_ZEN2HAN_DIGIT_COLON)
    nony_num = pd.to_numeric(nony.str.extract(r"(\d+)")[0], errors="coerce")
    df["NONYUHIBIN"] = nony_num.apply(lambda n: f"{int(n):02d}" if pd.notna(n) else "")
    df["入車時間"] = df["入車時間"].apply(_normalize_hhmm)
    return df


def parse_groupeddata_json(cell_text: str) -> list[dict]:
    """
    Excelセル内の GroupedData(JSON文字列 or エスケープ済み文字列) を list[dict] 化。
    """
    if isinstance(cell_text, list):
        return [x for x in cell_text if isinstance(x, dict)]
    if isinstance(cell_text, dict):
        return [cell_text]
    if cell_text is None:
        return []
    try:
        if pd.isna(cell_text):
            return []
    except Exception:
        pass
    raw = str(cell_text).strip()
    if not raw:
        return []
    candidates = [raw]
    if raw.startswith("\"") and raw.endswith("\"") and len(raw) >= 2:
        candidates.append(raw[1:-1])
    try:
        decoded_once = json.loads(raw)
        candidates.append(decoded_once)
    except Exception:
        pass
    for cand in candidates:
        obj = cand
        if isinstance(cand, str):
            try:
                obj = json.loads(cand)
            except Exception:
                continue
        if isinstance(obj, str):
            try:
                obj = json.loads(obj)
            except Exception:
                continue
        if isinstance(obj, list):
            return [item for item in obj if isinstance(item, dict)]
        if isinstance(obj, dict):
            return [obj]
    return []


def extract_vendor_and_order_suffix(items: list[dict]) -> tuple[str, str]:
    """
    GroupedDataの先頭要素から ベンダ名 と NONYUHIBIN末尾2桁 を取り出す。
    """
    if not items:
        return "", ""
    head = items[0] or {}
    vendor = str(head.get("OData_納入先") or head.get("OData__x7d0d__x5165__x5148_", "")).strip()
    nony = str(head.get("NONYUHIBIN", "")).strip().translate(_ZEN2HAN_DIGIT_COLON)
    order2 = nony[-2:] if len(nony) >= 2 else ""
    return vendor, order2


def attach_pickup_start_time(
    spo_df: pd.DataFrame,
    master_df: pd.DataFrame,
    unmatched_csv_path: Optional[Path] = None,
) -> pd.DataFrame:
    """
    GroupedDataから (OData_納入先, NONYUHIBIN末尾2桁) を抽出し、入車時間マスタを参照して引取開始時間を付与。
    
    武部の場合：同じ入車時間の便を1つのグループとして扱い、前グループの入車時間+10分を設定。
    （例：武部1-4便が同じ入車時間10:00の場合、前グループの入車時間+10分を設定）
    
    他の納入先：前便(便番号-1)の入車時間+10分を設定。
    （例：2便選択時 → 1便入車08:30+10分=08:40から開始）
    
    unmatched_csv_path が与えられた場合、未ヒットの index/vendor/order2 を CSV で出力。
    """
    if spo_df is None or spo_df.empty:
        return spo_df
    if master_df is None or master_df.empty:
        return spo_df
    for col in ("OData_納入先", "NONYUHIBIN", "入車時間"):
        if col not in master_df.columns:
            return spo_df
    master = master_df.copy()
    master["OData_納入先"] = master["OData_納入先"].astype(str).str.strip()
    master["NONYUHIBIN"] = master["NONYUHIBIN"].astype(str).str.strip().str.translate(_ZEN2HAN_DIGIT_COLON)
    master["入車時間"] = master["入車時間"].astype(str).str.strip()
    master = master[(master["OData_納入先"] != "") & (master["NONYUHIBIN"] != "")]
    master_map = {(r["OData_納入先"], r["NONYUHIBIN"]): r["入車時間"] for _, r in master.iterrows()}
    
    # 納入先ごとに入車時間でグループ化し、前グループの入車時間を取得できるマップを作成
    # vendor_time_groups[vendor] = [(入車時間(分), [便番号リスト]), ...] を入車時間昇順でソート
    vendor_time_groups: Dict[str, List[Tuple[int, List[str]]]] = {}
    for (v, bin_no), pickup_time in master_map.items():
        if v not in vendor_time_groups:
            vendor_time_groups[v] = {}
        normalized = _normalize_hhmm(pickup_time)
        if normalized:
            try:
                hh, mm = normalized.split(":", 1)
                mins = int(hh) * 60 + int(mm)
                if mins not in vendor_time_groups[v]:
                    vendor_time_groups[v][mins] = []
                vendor_time_groups[v][mins].append(bin_no)
            except:
                pass
    
    # 辞書を入車時間昇順のリストに変換
    vendor_sorted_groups: Dict[str, List[Tuple[int, List[str]]]] = {}
    for v, time_dict in vendor_time_groups.items():
        sorted_times = sorted(time_dict.keys())
        vendor_sorted_groups[v] = [(t, time_dict[t]) for t in sorted_times]
    
    def _to_minutes(hhmm: str) -> Optional[int]:
        s = _normalize_hhmm(hhmm)
        if not s:
            return None
        try:
            hh, mm = s.split(":", 1)
            return int(hh) * 60 + int(mm)
        except Exception:
            return None
    
    def _minutes_to_time(mins: int) -> str:
        """分をHH:MM形式に変換"""
        if mins < 0:
            mins = 0
        hh = mins // 60
        mm = mins % 60
        return f"{hh:02d}:{mm:02d}"
    
    def _get_prev_group_time(vendor: str, current_mins: int) -> Optional[int]:
        """
        指定された納入先と入車時間に対して、前グループの入車時間を取得。
        同じ入車時間の便はグループとして扱い、その前のグループの入車時間を返す。
        """
        if vendor not in vendor_sorted_groups:
            return None
        groups = vendor_sorted_groups[vendor]
        prev_time = None
        for time_mins, _ in groups:
            if time_mins >= current_mins:
                break
            prev_time = time_mins
        return prev_time

    out = spo_df.copy()
    unmatched_rows = []
    for idx, row in out.iterrows():
        items = parse_groupeddata_json(row.get("GroupedData", ""))

        # 混載対応: 全要素を照合し、最も遅い入車時間を採用
        best_time = ""
        best_min: Optional[int] = None
        seen_keys = set()
        for it in items:
            if not isinstance(it, dict):
                continue
            vendor = str(it.get("OData_納入先") or it.get("OData__x7d0d__x5165__x5148_", "")).strip()
            nony = str(it.get("NONYUHIBIN", "")).strip().translate(_ZEN2HAN_DIGIT_COLON)
            order2 = nony[-2:] if len(nony) >= 2 else ""
            if not vendor or not order2:
                continue
            key = (vendor, order2)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            pickup = master_map.get(key, "")
            if pickup:
                mins = _to_minutes(pickup)
                if mins is None:
                    continue
                if best_min is None or mins > best_min:
                    best_min = mins
                    # 武部の場合：同じ入車時間のグループの前グループの入車時間+10分
                    # 他の納入先：前便(便番号-1)の入車時間+10分
                    if vendor == "武部":
                        # 武部専用：同じ入車時間の便はグループとして扱う
                        prev_group_time = _get_prev_group_time(vendor, mins)
                        if prev_group_time is not None:
                            best_time = _minutes_to_time(prev_group_time + 10)
                        else:
                            # 前グループがない場合（最初のグループ）は当該便の入車時間+10分
                            best_time = _minutes_to_time(mins + 10)
                    else:
                        # 他の納入先：前便(便番号-1)の入車時間+10分
                        try:
                            current_bin = int(order2)
                            if current_bin > 1:
                                prev_bin = f"{current_bin - 1:02d}"
                                prev_pickup = master_map.get((vendor, prev_bin), "")
                                if prev_pickup:
                                    prev_mins = _to_minutes(prev_pickup)
                                    if prev_mins is not None:
                                        best_time = _minutes_to_time(prev_mins + 10)
                                        continue
                        except (ValueError, TypeError):
                            pass
                        # 前便が見つからない場合や1便の場合は当該便の入車時間+10分
                        best_time = _minutes_to_time(mins + 10)
            else:
                if unmatched_csv_path is not None:
                    unmatched_rows.append({"index": idx, "vendor": vendor, "order2": order2})

        existing = str(out.at[idx, "引取開始時間"]) if "引取開始時間" in out.columns else ""
        if best_time and (not existing or pd.isna(existing)):
            out.at[idx, "引取開始時間"] = best_time
    if unmatched_csv_path is not None and unmatched_rows:
        Path(unmatched_csv_path).parent.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(unmatched_rows, columns=["index", "vendor", "order2"]).to_csv(
            unmatched_csv_path, index=False, encoding="utf-8-sig"
        )
    return out


def adjust_pickup_time_for_same_bin(spo_df: pd.DataFrame, master_df: pd.DataFrame = None) -> pd.DataFrame:
    """
    同じ工程・同じ便番号で複数のグループ（山）がある場合、引取開始時間を順次調整する。
    各山の開始時間は前の山の終了後から連続して設定される。
    締め切り条件による逆算は行わない。
    
    便番号はGroupedDataから取得する（NONYUHIBIN末尾2桁）
    工程ごとに独立して処理するため、異なる工程では同じ引取開始時間が設定されることがある。
    """
    if spo_df is None or spo_df.empty:
        return spo_df
    
    out = spo_df.copy()
    
    def _time_to_seconds(hhmm: str) -> Optional[int]:
        """HH:MM形式を秒に変換"""
        s = _normalize_hhmm(hhmm)
        if not s:
            return None
        try:
            hh, mm = s.split(":", 1)
            return int(hh) * 3600 + int(mm) * 60
        except Exception:
            return None
    
    def _seconds_to_time(secs: int) -> str:
        """秒をHH:MM形式に変換"""
        if secs < 0:
            secs = 0
        hh = secs // 3600
        mm = (secs % 3600) // 60
        return f"{hh:02d}:{mm:02d}"
    
    # 休憩時間の定義（秒単位）
    # 1直: 8:30~8:40, 10:40~11:25, 12:55~13:05
    # 2直: 18:45~18:55, 20:55~21:40, 23:10~23:20
    BREAK_TIMES = [
        # 1直
        (8*3600 + 30*60, 8*3600 + 40*60),    # 8:30~8:40
        (10*3600 + 40*60, 11*3600 + 25*60),  # 10:40~11:25
        (12*3600 + 55*60, 13*3600 + 5*60),   # 12:55~13:05
        # 2直
        (18*3600 + 45*60, 18*3600 + 55*60),  # 18:45~18:55
        (20*3600 + 55*60, 21*3600 + 40*60),  # 20:55~21:40
        (23*3600 + 10*60, 23*3600 + 20*60),  # 23:10~23:20
    ]
    
    def _adjust_for_breaks(start_secs: int, work_duration: int) -> int:
        """
        作業開始時間と作業時間を受け取り、休憩時間を避けた開始時間を返す。
        作業時間帯が休憩時間と重なる場合、休憩後にずらす。
        """
        adjusted_start = start_secs
        for break_start, break_end in BREAK_TIMES:
            work_end = adjusted_start + work_duration
            # 作業開始が休憩中の場合 → 休憩終了後に開始
            if break_start <= adjusted_start < break_end:
                adjusted_start = break_end
            # 作業中に休憩が始まる場合 → 休憩終了後に開始
            elif adjusted_start < break_start < work_end:
                adjusted_start = break_end
        return adjusted_start
    
    def _extract_bin_info(grouped_data_str: str) -> tuple:
        """GroupedDataのJSONから便番号（NONYUHIBIN末尾2桁）と納入先を取得"""
        items = parse_groupeddata_json(grouped_data_str)
        if not items:
            return "", ""
        # 先頭要素から取得
        head = items[0] if isinstance(items[0], dict) else {}
        nony = str(head.get("NONYUHIBIN", "")).strip().translate(_ZEN2HAN_DIGIT_COLON)
        vendor = str(head.get("OData_納入先") or head.get("OData__x7d0d__x5165__x5148_", "")).strip()
        bin_num = nony[-2:] if len(nony) >= 2 else ""
        return bin_num, vendor
    
    # 入車時間マスタからマッピングを作成
    master_map = {}
    if master_df is not None and not master_df.empty:
        for col in ("OData_納入先", "NONYUHIBIN", "入車時間"):
            if col not in master_df.columns:
                break
        else:
            master = master_df.copy()
            master["OData_納入先"] = master["OData_納入先"].astype(str).str.strip()
            master["NONYUHIBIN"] = master["NONYUHIBIN"].astype(str).str.strip().str.translate(_ZEN2HAN_DIGIT_COLON)
            master["入車時間"] = master["入車時間"].astype(str).str.strip()
            master = master[(master["OData_納入先"] != "") & (master["NONYUHIBIN"] != "")]
            master_map = {(r["OData_納入先"], r["NONYUHIBIN"]): r["入車時間"] for _, r in master.iterrows()}
    
    # 便番号と納入先を各行に付与
    bin_info = out["GroupedData"].apply(_extract_bin_info)
    out["_便番号"] = bin_info.apply(lambda x: x[0])
    out["_納入先"] = bin_info.apply(lambda x: x[1])
    
    # 工程列が存在するか確認
    has_process_col = "工程" in out.columns
    
    # 工程ごとに引取開始時間が重複しないように調整
    # 同じ工程内では山を連続して処理する（前の山の終了後に次の山を開始）
    # 休憩時間を考慮して作業が休憩時間に被らないように調整
    if has_process_col:
        for proc, proc_group in out.groupby("工程", sort=False):
            # 引取開始時間でソートして処理
            proc_sorted = proc_group.sort_values(by="引取開始時間", key=lambda x: x.astype(str))
            
            prev_end_seconds = None
            for idx, row in proc_sorted.iterrows():
                start_time_str = str(row["引取開始時間"]).strip()
                start_seconds = _time_to_seconds(start_time_str)
                work_duration = int(row.get("引取工数", 0))
                
                if prev_end_seconds is not None:
                    # 前の山の終了後から開始（工程内で連続して作業）
                    start_seconds = prev_end_seconds
                
                # 休憩時間を考慮して調整
                if start_seconds is not None:
                    start_seconds = _adjust_for_breaks(start_seconds, work_duration)
                    out.at[idx, "引取開始時間"] = _seconds_to_time(start_seconds)
                    prev_end_seconds = start_seconds + work_duration
    else:
        # 工程列がない場合は便番号でグループ化
        for bin_num, group in out.groupby("_便番号", sort=False):
            if not bin_num or pd.isna(bin_num):
                continue
            
            group_sorted = group.sort_values(by="引取開始時間", key=lambda x: x.astype(str))
            
            prev_end_seconds = None
            for idx, row in group_sorted.iterrows():
                start_time_str = str(row["引取開始時間"]).strip()
                start_seconds = _time_to_seconds(start_time_str)
                work_duration = int(row.get("引取工数", 0))
                
                if prev_end_seconds is not None:
                    start_seconds = prev_end_seconds
                
                # 休憩時間を考慮して調整
                if start_seconds is not None:
                    start_seconds = _adjust_for_breaks(start_seconds, work_duration)
                    out.at[idx, "引取開始時間"] = _seconds_to_time(start_seconds)
                    prev_end_seconds = start_seconds + work_duration
    
    # 一時列を削除して返す
    out = out.drop(columns=["_便番号", "_納入先"])
    
    return out


# ===== SPO（SharePoint）アップロード想定の 1山=1行 =====
def build_spo_export_df(proc_details: pd.DataFrame, mountain_proc_map: dict) -> pd.DataFrame:
    """
    SharePointリストの列に合わせた 1山=1行 のDFを組み立てる
    """
    cols_out = [
        "タイトル", "工程", "groupdata", "GroupedData",
        "Max移動工数", "グループ番号", "パレット数", "引取工数",
        "引取開始時間", "id", "済", "実施者", "順番",
        "照合日", "照合済", "割込み作業名", "更新日時", "登録日時"
    ]
    if proc_details is None or proc_details.empty:
        return pd.DataFrame(columns=cols_out)
    df = proc_details.copy()
    df["移動工数"] = pd.to_numeric(df.get("移動工数", np.nan), errors="coerce")
    rows = []
    now_iso = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    
    # 工程ごとの山番号カウンターを作成
    proc_mountain_counter = {}  # {工程ラベル: カウンター}
    
    for yama, sub in df.groupby("山通番", sort=True):
        pal = int(sub.shape[0])
        max_cost = float(sub["移動工数"].max()) if sub["移動工数"].notna().any() else 0.0
        pick_cost = float(np.round(
            max_cost + BASE_ONE_TIME + ((pal - 1) * MIDDLE_WORK) + (pal * BASE_PER_PAL), 0
        ))
        gd_json = build_groupeddata_json_for_mountain(sub)
        y_proc = str(mountain_proc_map.get(int(yama), "4"))
        # 表示用ラベル（1工程/2工程/3工程/4工程）
        y_proc_label = f"{y_proc}工程" if y_proc in ("1","2","3") else "4工程"
        # 工程ごとの山番号をカウント
        if y_proc_label not in proc_mountain_counter:
            proc_mountain_counter[y_proc_label] = 0
        proc_mountain_counter[y_proc_label] += 1
        proc_mountain_num = proc_mountain_counter[y_proc_label]
        rows.append({
            "タイトル": f"山{proc_mountain_num}",
            "工程": y_proc_label,
            "groupdata": gd_json,
            "GroupedData": gd_json,
            "Max移動工数": max_cost,
            "グループ番号": int(yama),
            "パレット数": pal,
            "引取工数": int(pick_cost),
            "引取開始時間": "",
            "id": int(yama),
            "済": "",
            "実施者": "",
            "順番": 0,
            "照合日": "",
            "照合済": "",
            "割込み作業名": "",
            "更新日時": now_iso,
            "登録日時": now_iso,
        })
    out = pd.DataFrame(rows, columns=cols_out)
    # 型の明確化
    for c in ("Max移動工数", "グループ番号", "パレット数", "引取工数", "id", "順番"):
        if c in out.columns:
            if c == "Max移動工数":
                out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).astype(float)
            else:
                out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).astype(int)
    text_cols = ["タイトル", "工程", "groupdata", "GroupedData", "引取開始時間",
                 "済", "実施者", "照合日", "照合済", "割込み作業名", "更新日時", "登録日時"]
    out = _protect_excel_injection(out, text_cols)
    return out

def export_spo_xlsx(spo_df: pd.DataFrame, out_dir: str, base_name: str = "SPOアップロード用") -> str:
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    path = os.path.join(out_dir, f"{base_name}.xlsx")
    spo_df.to_excel(path, index=False, engine="openpyxl")
    _add_table_exact(path, "SPOExport")
    return path

def append_to_spo_history(spo_df: pd.DataFrame, out_dir: str, history_name: str = "SPOアップロード用_履歴") -> str:
    """SPOアップロード用のデータを履歴ファイルに追記する"""
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    history_path = os.path.join(out_dir, f"{history_name}.xlsx")
    
    # 出力日時列を追加
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    spo_with_time = spo_df.copy()
    spo_with_time.insert(0, "出力日時", timestamp)
    
    # 既存ファイルがあれば読み込んで追記、なければ新規作成
    if os.path.exists(history_path):
        try:
            existing_df = pd.read_excel(history_path, engine="openpyxl")
            # 既存データと新規データを結合
            combined_df = pd.concat([existing_df, spo_with_time], ignore_index=True)
        except Exception as e:
            print(f"既存履歴ファイルの読み込みに失敗: {e}")
            combined_df = spo_with_time
    else:
        combined_df = spo_with_time
    
    # 履歴ファイルに保存
    combined_df.to_excel(history_path, index=False, engine="openpyxl")
    _add_table_exact(history_path, "SPOHistory")
    
    return history_path

# ============ Tkinter GUI ============
class App(tb.Window):
    def __init__(self):
        # cosmo テーマで起動
        super().__init__(themename="cosmo")
        self.title("仕分け・工程割り振りビューア（ttkbootstrap: cosmo)")
        self.geometry("1280x860")
        # ttkbootstrap のスタイル（読み取り専用の self.style をそのまま使う）
        self.colors = self.style.colors
        # 念のため、darkly を強制適用
        try:
            if getattr(self.style, "theme", None) and self.style.theme.name != "cosmo":
                self.style.theme_use("cosmo")
        except Exception:
            self.style.theme_use("cosmo")
        # フィールド初期化
        self.height_cap = tk.IntVar(value=DEFAULT_HEIGHT_CAP)
        self.mixing_key = tk.StringVar(value=DEFAULT_MIXING_KEY)
        self.selections = []
        self.filtered = pd.DataFrame()
        self.expanded = pd.DataFrame()
        self.group_results = {}
        self.group_details = {}
        self.size1_mixed_summary = None
        self.size1_mixed_details = None
        self.process_map = {}
        self.proc_details = pd.DataFrame()
        self.proc_summary = pd.DataFrame()
        self.mountain_proc = pd.DataFrame()
        self.mountain_proc_map = {}
        self.all_mountain_details = pd.DataFrame()
        self.auto_export_csv = True
        self.export_dir = r"C:\Users\1588386\トヨタ自動車株式会社\資材物流課　車体部　連絡・共有 - テスト_河崎"
        self.export_encoding = "utf-8-sig"
        # 入車時間マスタ管理
        self.master_data = pd.DataFrame(columns=["OData_納入先", "NONYUHIBIN", "入車時間"])
        # UI 初期化
        self.build_ui()
        # テーマ色タグ再設定
        self.reapply_treeview_tags()
        # ルール読み込みとUI反映
        self.load_process_rules()
        self.refresh_proc_rule_tree()
        print("ttkbootstrap theme =>", self.style.theme.name)
        print("ttkbootstrap theme =>", self.style.theme.name)  # 'cosmo'確認
        # Enterキーで実行
        self.bind("<Return>", lambda e: self.run())
        self.bind("<KP_Enter>", lambda e: self.run())
    def reapply_treeview_tags(self):
        """Treeviewタグをテーマ色に合わせて再設定（古いttkbootstrap対応版）"""
        # 淡色は固定コードで指定
        light_primary = "#DFF0FF"   # 薄い青
        light_success = "#E8F7DF"   # 薄い緑
        light_danger  = "#FBE1EF"   # 薄いピンク
        light_warning = "#FFF6BF"   # 薄い黄色
        light_muted   = "#E0E0E0"   # 薄いグレー
        # 工程別サマリ
        try:
            self.kb_summary.tag_configure("proc_1", background=light_primary)
            self.kb_summary.tag_configure("proc_2", background=light_success)
            self.kb_summary.tag_configure("proc_3", background=light_danger)
            self.kb_summary.tag_configure("proc_unset", background=light_warning)
        except Exception:
            pass
        # 工程明細
        try:
            self.kb_detail.tag_configure("proc_unset", background=light_warning)
            self.kb_detail.tag_configure("proc_mismatch", background=light_muted)
        except Exception:
            pass
        # 混載・納入先関連
        try:
            self.basic_summary.tag_configure("basic_mixed", background=light_warning)
            self.mix_summary.tag_configure("mixed_true", background=light_warning)
            self.dest_summary.tag_configure("dst_mixed", background=light_warning)
        except Exception:
            pass
    # ルールファイルパス
    def get_rule_path(self) -> Path:
        home = Path.home()
        base_dir = next((p for p in [home / "Documents", home / "ドキュメント"] if p.exists()), home / "Documents")
        return base_dir / PROC_RULE_FILENAME
    def load_process_rules(self):
        path = self.get_rule_path()
        self.process_map = {}
        if path.exists():
            try:
                df = read_csv_ja(path)
                if "納入先" in df.columns and "工程" in df.columns:
                    for _, r in df.iterrows():
                        dest = str(r["納入先"]).strip()
                        try:
                            proc = int(r["工程"])
                        except Exception:
                            continue
                        if dest and proc in (1, 2, 3):
                            self.process_map[dest] = proc
            except Exception as e:
                messagebox.showwarning("工程ルール読込", f"工程割当ルールの読込に失敗しました: {e}")
    def save_process_rules(self):
        path = self.get_rule_path()
        if not self.process_map:
            if path.exists():
                try:
                    path.unlink()
                except Exception:
                    pass
            messagebox.showinfo("工程ルール保存", "現在ルールが空のため、ファイルは作成/更新されていません。")
            return
        df = pd.DataFrame([{"納入先": k, "工程": v} for k, v in sorted(self.process_map.items())])
        try:
            write_csv_ja(df, path)
            messagebox.showinfo("工程ルール保存", f"保存しました: {path}")
        except Exception as e:
            messagebox.showerror("工程ルール保存", f"保存に失敗しました: {e}")
    def build_ui(self):
        # 左ペイン
        left = tb.Frame(self)  # ttk.Frame -> tb.Frame にすると枠色もテーマ統一
        left.pack(side="left", fill="y", padx=6, pady=6)
        tb.Label(left, text="便名（複数選択可）", bootstyle=PRIMARY).pack(anchor="w")
        self.route_list = tk.Listbox(left, selectmode="extended", height=8, exportselection=False)
        self.route_list.pack(fill="x")
        self.route_list.bind("<<ListboxSelect>>", lambda e: self.refresh_candidates())
        self.summary_mode = tk.BooleanVar(value=True)
        tb.Checkbutton(left, text="受入をまとめて展開（便名＋オーダー）",
               variable=self.summary_mode,
               command=self.refresh_candidates,
               bootstyle=INFO).pack(anchor="w", pady=(4,8))
        tb.Label(left, text="受入（まとめOFF時に選択）", bootstyle=SECONDARY).pack(anchor="w")
        self.receipt_list = tk.Listbox(left, selectmode="extended", height=8, exportselection=False)
        self.receipt_list.pack(fill="x")
        self.receipt_list.bind("<<ListboxSelect>>", lambda e: self.refresh_orders_for_receipt())
        tb.Label(left, text="オーダー", bootstyle=SECONDARY).pack(anchor="w")
        self.order_list = tk.Listbox(left, selectmode="extended", height=8, exportselection=False)
        self.order_list.pack(fill="x")
        # ダブルクリックで追加
        self.order_list.bind("<Double-Button-1>", lambda e: self.add_selection())
        btns = ttk.Frame(left); btns.pack(fill="x", pady=6)
        tb.Button(btns, text="候補更新", command=self.refresh_candidates, bootstyle=INFO).pack(side="left", padx=2)
        tb.Button(btns, text="追加", command=self.add_selection, bootstyle=SUCCESS).pack(side="left", padx=2)
        tb.Button(btns, text="削除", command=self.delete_selection, bootstyle=DANGER).pack(side="left", padx=2)
        tb.Button(btns, text="クリア", command=self.clear_selection, bootstyle=SECONDARY).pack(side="left", padx=2)
        tb.Button(btns, text="実行", command=self.run, bootstyle=PRIMARY).pack(side="left", padx=6)
        ttk.Label(left, text="選択一覧（便名×受入×オーダー）").pack(anchor="w")
        self.sel_tree = ttk.Treeview(left, columns=("便名","受入","オーダー"), show="headings", height=8)
        for c in ("便名","受入","オーダー"):
            self.sel_tree.heading(c, text=c); self.sel_tree.column(c, width=120, anchor="w")
        self.sel_tree.pack(fill="x", pady=(2,8))
        # Deleteキーでも削除可能に
        self.sel_tree.bind("<Delete>", lambda e: self.delete_selection())
        cfg = ttk.LabelFrame(left, text="設定")
        cfg.pack(fill="x", pady=6)
        ttk.Label(cfg, text="混載キー").grid(row=0, column=0, sticky="w")
        self.mixing_combo = ttk.Combobox(cfg, values=["UKEIRE","納入先コード","SYUKKASAKI","NONYUHIBIN","納入先"],
                                        textvariable=self.mixing_key, state="readonly")
        self.mixing_combo.grid(row=0, column=1, sticky="ew", padx=4)
        ttk.Label(cfg, text="高さ上限").grid(row=1, column=0, sticky="w")
        ttk.Spinbox(cfg, from_=1000, to=5000, increment=50, textvariable=self.height_cap, width=8).grid(row=1, column=1, sticky="w", padx=4)
        cfg.columnconfigure(1, weight=1)
        # 右ペイン：Notebook
        right = ttk.Notebook(self); right.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        self.notebook = right
        # 基本グループタブ
        tab_basic = ttk.Frame(right); right.add(tab_basic, text="基本グループ")
        self.basic_summary = ttk.Treeview(
            tab_basic,
            columns=("サイズ種類","グループ番号","パレット数","Max移動工数","引取工数","高さ合計","納入先一覧","混載"),
            show="headings"
        )
        for c in self.basic_summary["columns"]:
            self.basic_summary.heading(c, text=c)
            self.basic_summary.column(c, width=120 if c not in ("納入先一覧","混載") else (360 if c=="納入先一覧" else 60), anchor="w")
        self.basic_summary.pack(fill="both", expand=True)
        self.basic_summary.bind("<<TreeviewSelect>>", self.on_basic_select)
        ttk.Label(tab_basic, text="明細").pack(anchor="w")
        self.basic_detail = ttk.Treeview(tab_basic, show="headings"); self.basic_detail.pack(fill="both", expand=True)
        # 混載=赤色タグ
        self.basic_summary.tag_configure("basic_mixed", background="#FFC9C9")
        # 種類1 混載タブ（total工数列を追加）
        tab_mix = ttk.Frame(right); right.add(tab_mix, text="種類1 混載")
        self.tab_mix = tab_mix
        self.mix_summary = ttk.Treeview(
            tab_mix,
            columns=("混載","山通番","パレット数","Max移動工数","引取工数","total工数","高さ合計","混載キー種類数","混載フラグ","混載キー一覧"),
            show="headings"
        )
        for c in self.mix_summary["columns"]:
            self.mix_summary.heading(c, text=c)
            if c == "混載":
                self.mix_summary.column(c, width=60, anchor="w")
            elif c == "total工数":
                self.mix_summary.column(c, width=140, anchor="w")
            else:
                self.mix_summary.column(c, width=120, anchor="w")
        self.mix_summary.pack(fill="both", expand=True)
        self.mix_summary.bind("<<TreeviewSelect>>", self.on_mix_select)
        ttk.Label(tab_mix, text="明細").pack(anchor="w")
        self.mix_detail = ttk.Treeview(tab_mix, show="headings"); self.mix_detail.pack(fill="both", expand=True)
        # 混載=赤色タグ
        self.mix_summary.tag_configure("mixed_true", background="#FFC9C9")
        # 山別 納入先タブ
        tab_dst = ttk.Frame(right); right.add(tab_dst, text="山別 納入先")
        self.dest_summary = ttk.Treeview(
            tab_dst,
            columns=("山通番","納入先数","納入先一覧","パレット数","高さ合計"),
            show="headings"
        )
        for c in self.dest_summary["columns"]:
            self.dest_summary.heading(c, text=c)
            self.dest_summary.column(c, width=140 if c not in ("納入先一覧",) else 400, anchor="w")
        self.dest_summary.pack(fill="both", expand=True)
        self.dest_summary.bind("<<TreeviewSelect>>", self.on_dest_select)
        ttk.Label(tab_dst, text="明細（山内の納入先別）").pack(anchor="w")
        self.dest_detail = ttk.Treeview(tab_dst, show="headings"); self.dest_detail.pack(fill="both", expand=True)
        # 納入先数>=2=赤色タグ
        self.dest_summary.tag_configure("dst_mixed", background="#FFC9C9")
        # 工程設定タブ
        tab_proc = ttk.Frame(right); right.add(tab_proc, text="工程設定")
        self.tab_proc = tab_proc
        # 上段：納入先候補 + 検索
        filt_frame = ttk.Frame(tab_proc); filt_frame.pack(fill="x", padx=6, pady=(6,2))
        ttk.Label(filt_frame, text="納入先フィルタ").pack(side="left")
        self.proc_filter = tk.StringVar(value="")
        ent = ttk.Entry(filt_frame, textvariable=self.proc_filter, width=30)
        ent.pack(side="left", padx=6)
        ent.bind("<KeyRelease>", lambda e: self.update_proc_candidates())
        assign_frame = ttk.Frame(tab_proc); assign_frame.pack(fill="both", expand=True, padx=6, pady=4)
        # 候補リスト
        left_cand = ttk.LabelFrame(assign_frame, text="納入先（候補）")
        left_cand.pack(side="left", fill="both", expand=True, padx=(0,4))
        self.proc_cand_tree = ttk.Treeview(left_cand, columns=("納入先",), show="headings", height=12, selectmode="extended")
        self.proc_cand_tree.heading("納入先", text="納入先")
        self.proc_cand_tree.column("納入先", width=260, anchor="w")
        self.proc_cand_tree.pack(fill="both", expand=True, padx=4, pady=4)
        # 中央：割当操作
        mid_ops = ttk.LabelFrame(assign_frame, text="割当操作")
        mid_ops.pack(side="left", fill="y", padx=4)
        ttk.Label(mid_ops, text="工程").pack(anchor="w", padx=6, pady=(8,2))
        self.proc_value = tk.StringVar(value="1")
        self.proc_combo = ttk.Combobox(mid_ops, values=["1","2","3"], textvariable=self.proc_value, state="readonly", width=6)
        self.proc_combo.pack(anchor="w", padx=6)
        ttk.Button(mid_ops, text="割当/更新", command=self.assign_process_to_selected).pack(anchor="w", padx=6, pady=(10,4))
        ttk.Button(mid_ops, text="選択解除", command=lambda: self.proc_cand_tree.selection_remove(*self.proc_cand_tree.selection())).pack(anchor="w", padx=6, pady=2)
        ttk.Separator(mid_ops, orient="horizontal").pack(fill="x", padx=6, pady=10)
        ttk.Button(mid_ops, text="選択ルール削除", command=self.delete_process_rules_selected).pack(anchor="w", padx=6, pady=(2,2))
        ttk.Button(mid_ops, text="ルール保存", command=self.save_process_rules).pack(anchor="w", padx=6, pady=(10,2))
        # 右：現在のルール
        right_rules = ttk.LabelFrame(assign_frame, text="現在の割当ルール（納入先→工程）")
        right_rules.pack(side="left", fill="both", expand=True, padx=(4,0))
        self.proc_rule_tree = ttk.Treeview(right_rules, columns=("納入先","工程"), show="headings", height=12, selectmode="extended")
        for c in ("納入先","工程"):
            self.proc_rule_tree.heading(c, text=c)
            self.proc_rule_tree.column(c, width=120 if c=="工程" else 260, anchor="w")
        self.proc_rule_tree.pack(fill="both", expand=True, padx=4, pady=4)
        # 工程別かんばんタブ（山工程列＋色タグ）
        tab_kb = ttk.Frame(right); right.add(tab_kb, text="工程別かんばん")
        self.tab_kb = tab_kb
        self.kb_summary = ttk.Treeview(
            tab_kb,
            columns=("山通番","山工程","工程1","工程2","工程3","4工程","合計"),
            show="headings",
            height=12
        )
        for c in self.kb_summary["columns"]:
            w = 120 if c != "山工程" else 90
            self.kb_summary.heading(c, text=c)
            self.kb_summary.column(c, width=w, anchor="w")
        self.kb_summary.pack(fill="both", expand=True, padx=6, pady=(6,2))
        self.kb_summary.tag_configure("proc_1", background="#DFF0FF")
        self.kb_summary.tag_configure("proc_2", background="#E8F7DF")
        self.kb_summary.tag_configure("proc_3", background="#FBE1EF")
        self.kb_summary.tag_configure("proc_4", background="#FFF6BF")
        self.kb_summary.bind("<<TreeviewSelect>>", self.on_kb_select)
        self.kb_yama_label = ttk.Label(tab_kb, text="選択中の山：- / 山工程：-")
        self.kb_yama_label.pack(anchor="w", padx=6)
        ttk.Label(tab_kb, text="明細（工程・工程内No）").pack(anchor="w", padx=6)
        self.kb_detail = ttk.Treeview(tab_kb, show="headings", height=14)
        self.kb_detail.pack(fill="both", expand=True, padx=6, pady=(2,6))
        self.kb_detail.tag_configure("proc_unset", background="#FFF6BF")
        self.kb_detail.tag_configure("proc_mismatch", background="#FFE0CC")
        # Total工数タブ
        tab_total = ttk.Frame(right); right.add(tab_total, text="Total工数")
        self.tab_total = tab_total
        self.total_tree = ttk.Treeview(
            tab_total,
            columns=("区分","キー","パレット数","total工数","納入先一覧"),
            show="headings",
            height=16
        )
        for c in self.total_tree["columns"]:
            w = 120
            if c in ("キー","納入先一覧"): w = 220
            if c == "total工数": w = 140
            self.total_tree.heading(c, text=c)
            self.total_tree.column(c, width=w, anchor="w")
        self.total_tree.pack(fill="both", expand=True, padx=6, pady=6)
        # 入車時間マスタタブ
        tab_master = ttk.Frame(right); right.add(tab_master, text="入車時間マスタ")
        self.tab_master = tab_master
        # 上部：ボタン
        master_btn_frame = ttk.Frame(tab_master)
        master_btn_frame.pack(fill="x", padx=6, pady=6)
        ttk.Button(master_btn_frame, text="マスタ読込", command=self.load_master).pack(side="left", padx=2)
        ttk.Button(master_btn_frame, text="マスタ保存", command=self.save_master).pack(side="left", padx=2)
        ttk.Button(master_btn_frame, text="行追加", command=self.add_master_row).pack(side="left", padx=2)
        ttk.Button(master_btn_frame, text="選択行削除", command=self.delete_master_row).pack(side="left", padx=2)
        ttk.Button(master_btn_frame, text="全クリア", command=self.clear_master).pack(side="left", padx=2)
        # 中央：データグリッド
        self.master_tree = ttk.Treeview(
            tab_master,
            columns=("OData_納入先", "NONYUHIBIN", "入車時間"),
            show="headings",
            height=20
        )
        for c in self.master_tree["columns"]:
            self.master_tree.heading(c, text=c)
            w = 250 if c == "OData_納入先" else 150
            self.master_tree.column(c, width=w, anchor="w")
        self.master_tree.pack(fill="both", expand=True, padx=6, pady=(0, 6))
        self.master_tree.bind("<Double-1>", self.edit_master_row)
        # 初期候補
        self.refresh_routes()
        self.update_proc_candidates()
        self.refresh_proc_rule_tree()
        # 入車時間マスタを読み込み
        self.load_master()
        # --- 追加部分（不要タブ削除） ---
        # Notebookの既存タブから必要なものだけ残す
        for tab_id in self.notebook.tabs():
            tab_text = self.notebook.tab(tab_id, option="text")
            if tab_text not in ("工程設定", "工程別かんばん"):
                self.notebook.forget(tab_id)
    # 候補更新（左ペイン）
    def refresh_routes(self):
        self.route_list.delete(0, "end")
        for r in get_routes():
            self.route_list.insert("end", r)
        self.refresh_candidates()
    def refresh_candidates(self):
        routes = [self.route_list.get(i) for i in self.route_list.curselection()] or []
        print(f"DEBUG: selected routes = {routes}")  # デバッグ用
        
        self.receipt_list.delete(0, "end")
        self.order_list.delete(0, "end")
        if not routes:
            return
        if not self.summary_mode.get():
            receipts_all = set()
            for route in routes:
                receipts_route = get_receipts_for_route(route)
                print(f"DEBUG: route '{route}' -> receipts = {receipts_route}")  # デバッグ用
                receipts_all.update(receipts_route)
            print(f"DEBUG: all receipts = {sorted(receipts_all)}")  # デバッグ用
            for rc in sorted(receipts_all):
                self.receipt_list.insert("end", rc)
        orders_all = set()
        if self.summary_mode.get():
            for route in routes:
                orders_route = get_orders_for_route(route)
                print(f"DEBUG: route '{route}' -> orders = {orders_route}")  # デバッグ用
                orders_all.update(orders_route)
        else:
            receipts = [self.receipt_list.get(i) for i in self.receipt_list.curselection()]
            if not receipts:
                receipts = [self.receipt_list.get(i) for i in range(self.receipt_list.size())]
            for route in routes:
                for rc in receipts:
                    orders_for_receipt = get_orders_for_route_receipt(route, rc)
                    print(f"DEBUG: route '{route}', receipt '{rc}' -> orders = {orders_for_receipt}")  # デバッグ用
                    orders_all.update(orders_for_receipt)
        print(f"DEBUG: all orders = {sorted(orders_all, reverse=True)}")  # デバッグ用    
        for od in sorted(orders_all, reverse=True):
            self.order_list.insert("end", od)
    
    def refresh_orders_for_receipt(self):
        """受入選択後にオーダーリストを更新（まとめ選択OFFの場合のみ有効）"""
        if self.summary_mode.get():
            return  # まとめ選択ONの場合は何もしない
        
        routes = [self.route_list.get(i) for i in self.route_list.curselection()] or []
        if not routes:
            return
        
        receipts = [self.receipt_list.get(i) for i in self.receipt_list.curselection()]
        if not receipts:
            return  # 受入が選択されていない場合は何もしない
        
        self.order_list.delete(0, "end")
        orders_all = set()
        for route in routes:
            for rc in receipts:
                orders_for_receipt = get_orders_for_route_receipt(route, rc)
                print(f"DEBUG: refresh_orders_for_receipt - route '{route}', receipt '{rc}' -> orders = {len(orders_for_receipt)}")
                orders_all.update(orders_for_receipt)
        
        for od in sorted(orders_all, reverse=True):
            self.order_list.insert("end", od)
        print(f"DEBUG: refresh_orders_for_receipt - total orders = {len(orders_all)}")
    
    # selections 操作
    def add_selection(self):
        routes = [self.route_list.get(i) for i in self.route_list.curselection()]
        orders = [self.order_list.get(i) for i in self.order_list.curselection()]
        if not routes or not orders:
            messagebox.showinfo("追加", "便名とオーダーを選択してください。（受入はまとめ選択OFF時のみ必要）")
            return
        new_items = []
        if self.summary_mode.get():
            for route in routes:
                for od in orders:
                    receipts = get_receipts_for_route_order(route, od)
                    if not receipts:
                        messagebox.showwarning("追加", f"便名={route}, オーダー={od} に該当する受入がありません。")
                        continue
                    for rc in receipts:
                        new_items.append({"便名": route, "受入": rc, "オーダー": od})
        else:
            receipts = [self.receipt_list.get(i) for i in self.receipt_list.curselection()]
            if not receipts:
                messagebox.showinfo("追加", "受入を選択してください。")
                return
            for route in routes:
                for rc in receipts:
                    for od in orders:
                        new_items.append({"便名": route, "受入": rc, "オーダー": od})
        uniq = {(s["便名"], s["受入"], s["オーダー"]) for s in (self.selections + new_items)}
        self.selections = [{"便名": a, "受入": b, "オーダー": c} for (a, b, c) in sorted(uniq)]
        self.refresh_selection_tree()
    def delete_selection(self):
        self.sel_tree.focus_set()
        sel_iids = list(self.sel_tree.selection())
        if not sel_iids:
            messagebox.showinfo("削除", "選択一覧で削除対象を選んでください。")
            return
        keys_to_remove = set()
        for iid in sel_iids:
            vals = self.sel_tree.item(iid, "values")
            if len(vals) >= 3:
                k = (str(vals[0]).strip(), str(vals[1]).strip(), str(vals[2]).strip())
                keys_to_remove.add(k)
        self.selections = [
            s for s in self.selections
            if (str(s["便名"]).strip(), str(s["受入"]).strip(), str(s["オーダー"]).strip()) not in keys_to_remove
        ]
        for iid in sel_iids:
            try:
                self.sel_tree.delete(iid)
            except Exception:
                pass
    def clear_selection(self):
        self.selections = []
        self.refresh_selection_tree()
    def refresh_selection_tree(self):
        for iid in self.sel_tree.get_children():
            self.sel_tree.delete(iid)
        for s in self.selections:
            a = str(s["便名"]).strip()
            b = str(s["受入"]).strip()
            c = str(s["オーダー"]).strip()
            iid = f"{a}|{b}|{c}"
            self.sel_tree.insert("", "end", iid=iid, values=(a, b, c))
    # 実行
    def run(self):
        if not self.selections:
            messagebox.showinfo("実行", "選択が空です。便名・受入・オーダーを追加してください。")
            return
        try:
            filtered, expanded, group_results, group_details, size1_mixed_summary, size1_mixed_details = run_pipeline(
                self.selections, self.height_cap.get(), self.mixing_key.get()
            )
        except Exception as e:
            messagebox.showerror("実行エラー", str(e))
            return
        self.filtered = filtered
        self.expanded = expanded
        self.group_results = group_results
        self.group_details = group_details
        self.size1_mixed_summary = size1_mixed_summary
        self.size1_mixed_details = size1_mixed_details
        # 工程割当の再計算
        self.recompute_process_assignment()
        self.update_basic_views()
        self.update_mix_views()  # 種類1 混載（total工数が表示されます）
        self.update_dest_views()
        self.update_proc_candidates()
        self.refresh_proc_rule_tree()
        self.update_kb_views()
        self.update_total_views()  # 追加：Total工数タブ
        # === 自動Excel出力（SPOのみ） ===
        try:
            if getattr(self, "auto_export_csv", False):
                spo_df = build_spo_export_df(self.proc_details, self.mountain_proc_map)
                master_df = pd.DataFrame()
                master_err = None
                try:
                    master_path = Path(__file__).with_name("入車時間マスタ.xlsx")
                    master_df = load_pickup_time_master_xlsx(master_path)
                    unmatched_path = Path(getattr(self, "export_dir", "exports")) / "SPOアップロード用_未ヒット一覧.csv"
                    spo_df = attach_pickup_start_time(spo_df, master_df, unmatched_csv_path=unmatched_path)
                    # 同じ便で複数グループがある場合の引取開始時間を逆算して調整
                    spo_df = adjust_pickup_time_for_same_bin(spo_df, master_df)
                except Exception as e:
                    master_err = str(e)
                if spo_df is not None and not spo_df.empty:
                    spo_path = export_spo_xlsx(spo_df, out_dir=getattr(self, "export_dir", "exports"),
                                               base_name="SPOアップロード用")
                    
                    # 履歴ファイルにも追記
                    try:
                        history_path = append_to_spo_history(spo_df, out_dir=getattr(self, "export_dir", "exports"))
                        print(f"履歴ファイルに追記しました: {history_path}")
                    except Exception as hist_err:
                        print(f"履歴ファイルへの追記に失敗: {hist_err}")
                    
                    try:
                        msg = f"SPO用Excelを出力しました。\n{spo_path}"
                        if master_err:
                            msg += f"\n\n入車時間の付与に失敗しました: {master_err}"
                        messagebox.showinfo("SPO出力", msg)
                        if os.name == "nt":
                            os.startfile(getattr(self, "export_dir", "exports"))
                    except Exception:
                        pass
        except Exception as e:
            messagebox.showwarning("Excel出力", f"SPO用Excel出力で警告/エラー: {e}")
        messagebox.showinfo("完了", "集計が完了しました。タブで結果をご確認ください。")
    # 表示更新（基本：全山＋納入先一覧＋混載★）
    def update_basic_views(self):
        for iid in self.basic_summary.get_children():
            self.basic_summary.delete(iid)
        df = compute_basic_groups(self.group_details, self.group_results, self.height_cap.get())
        for _, row in df.iterrows():
            values = [row[c] for c in self.basic_summary["columns"]]
            iid = f"{row['サイズ種類']}:{row['グループ番号']}"
            tags = ("basic_mixed",) if str(row.get("混載", "")) == "★" else ()
            self.basic_summary.insert("", "end", iid=iid, values=values, tags=tags)
        # 明細クリア
        self.basic_detail.configure(columns=())
        for iid in self.basic_detail.get_children():
            self.basic_detail.delete(iid)
    def on_basic_select(self, event=None):
        sel = self.basic_summary.selection()
        if not sel:
            return
        iid = sel[0]
        try:
            stype, gno = iid.split(":")
            gno = int(gno)
        except:
            return
        df = build_basic_detail_df(self.group_details, stype, gno)
        for iid in self.basic_detail.get_children():
            self.basic_detail.delete(iid)
        cols = list(df.columns)
        self.basic_detail.configure(columns=cols, show="headings")
        for c in cols:
            self.basic_detail.heading(c, text=c); self.basic_detail.column(c, width=120, anchor="w")
        for i, (_, row) in enumerate(df.iterrows(), 1):
            self.basic_detail.insert("", "end", iid=f"b:{i}", values=[row.get(c, "") for c in cols])
    # 表示更新（種類1 混載）
    def update_mix_views(self):
        for iid in self.mix_summary.get_children():
            self.mix_summary.delete(iid)
        df = compute_mixed_groups(self.size1_mixed_summary, self.size1_mixed_details, self.height_cap.get())
        for _, row in df.iterrows():
            is_mixed = bool(row.get("混載フラグ", False))
            star = "★" if is_mixed else ""
            values = [star] + [row[c] for c in self.mix_summary["columns"] if c != "混載"]
            tags = ("mixed_true",) if is_mixed else ()
            iid = f"{row['山通番']}"
            self.mix_summary.insert("", "end", iid=iid, values=values, tags=tags)
        # 明細クリア
        self.mix_detail.configure(columns=())
        for iid in self.mix_detail.get_children():
            self.mix_detail.delete(iid)
    def on_mix_select(self, event=None):
        sel = self.mix_summary.selection()
        if not sel:
            return
        yama = int(self.mix_summary.item(sel[0], "values")[1])
        df = build_mixed_detail_df(self.size1_mixed_details, yama)
        for iid in self.mix_detail.get_children():
            self.mix_detail.delete(iid)
        cols = list(df.columns)
        self.mix_detail.configure(columns=cols, show="headings")
        for c in cols:
            self.mix_detail.heading(c, text=c); self.mix_detail.column(c, width=120, anchor="w")
        for i, (_, row) in enumerate(df.iterrows(), 1):
            self.mix_detail.insert("", "end", iid=f"m:{i}", values=[row.get(c, "") for c in cols])
    # 表示更新（山別 納入先）
    def update_dest_views(self):
        for iid in self.dest_summary.get_children():
            self.dest_summary.delete(iid)
        df = compute_dest_by_mountain(self.size1_mixed_details, self.size1_mixed_summary, self.height_cap.get())
        for _, row in df.iterrows():
            values = [row[c] for c in self.dest_summary["columns"]]
            tag = ("dst_mixed",) if int(row.get("納入先数", 0)) >= 2 else ()
            iid = f"dst:{row['山通番']}"
            self.dest_summary.insert("", "end", iid=iid, values=values, tags=tag)
        # 明細クリア
        self.dest_detail.configure(columns=())
        for iid in self.dest_detail.get_children():
            self.dest_detail.delete(iid)
    def on_dest_select(self, event=None):
        sel = self.dest_summary.selection()
        if not sel:
            return
        vals = self.dest_summary.item(sel[0], "values")
        if not vals:
            return
        yama = int(vals[0])  # 1列目が山通番
        df = build_dest_detail_df(self.size1_mixed_details, yama)
        for iid in self.dest_detail.get_children():
            self.dest_detail.delete(iid)
        cols = list(df.columns) if not df.empty else ["納入先","パレット数","高さ合計","Max移動工数"]
        self.dest_detail.configure(columns=cols, show="headings")
        for c in cols:
            self.dest_detail.heading(c, text=c); self.dest_detail.column(c, width=140, anchor="w")
        for i, (_, row) in enumerate(df.iterrows(), 1):
            self.dest_detail.insert("", "end", iid=f"d:{i}", values=[row.get(c, "") for c in cols])
    # 工程タブ：候補更新（納入先名称のみ表示・数字だけを除外）
    def update_proc_candidates(self):
        def extract_dest_names_only(df: pd.DataFrame) -> list:
            if df is None or df.empty or ("納入先" not in df.columns):
                return []
            s = df["納入先"].astype(str).map(_normalize_dest_name)
            s = s[(s != "") & (~s.str.fullmatch(r"\d+"))]
            return sorted(set(s.tolist()))
        candidates = set()
        candidates.update(extract_dest_names_only(self.expanded))
        candidates.update(extract_dest_names_only(self.filtered))
        candidates.update(extract_dest_names_only(self.size1_mixed_details))
        candidates.update(extract_dest_names_only(self.all_mountain_details))
        candidates.update(extract_dest_names_only(df_shipments))
        all_dests = sorted(candidates)
        q = self.proc_filter.get().strip()
        if q:
            all_dests = [d for d in all_dests if q in d]
        all_dests = [d for d in all_dests if d not in self.process_map]
        for iid in self.proc_cand_tree.get_children():
            self.proc_cand_tree.delete(iid)
        for i, d in enumerate(all_dests, 1):
            self.proc_cand_tree.insert("", "end", iid=f"cand:{i}", values=(d,))
    def refresh_proc_rule_tree(self):
        for iid in self.proc_rule_tree.get_children():
            self.proc_rule_tree.delete(iid)
        for i, (dest, proc) in enumerate(sorted(self.process_map.items(), key=lambda x: (x[1], x[0])), 1):
            self.proc_rule_tree.insert("", "end", iid=f"rule:{i}", values=(dest, proc))
    def assign_process_to_selected(self):
        sel_iids = list(self.proc_cand_tree.selection())
        if not sel_iids:
            messagebox.showinfo("工程割当", "納入先（候補）を選択してください。Ctrl/Shiftで複数選択できます。")
            return
        try:
            proc = int(self.proc_value.get())
        except Exception:
            messagebox.showinfo("工程割当", "工程は 1〜3 を選択してください。")
            return
        for iid in sel_iids:
            vals = self.proc_cand_tree.item(iid, "values")
            if vals:
                dest = str(vals[0]).strip()
                if dest:
                    self.process_map[dest] = proc
        self.update_proc_candidates()
        self.refresh_proc_rule_tree()
        self.recompute_process_assignment()
        self.update_kb_views()
    def delete_process_rules_selected(self):
        sel_iids = list(self.proc_rule_tree.selection())
        if not sel_iids:
            messagebox.showinfo("ルール削除", "右側のルール表から削除対象を選択してください。")
            return
        for iid in sel_iids:
            vals = self.proc_rule_tree.item(iid, "values")
            if not vals:
                continue
            dest = str(vals[0]).strip()
            if dest in self.process_map:
                del self.process_map[dest]
        self.refresh_proc_rule_tree()
        self.update_proc_candidates()
        self.recompute_process_assignment()
        self.update_kb_views()
# 置換：工程割当（全サイズの山を対象に、入車時間ベースで自動割り振り）
    def recompute_process_assignment(self):
        """
        工程割当を入車時間マスタに基づいて自動で行う。
        - 種類1混載山 + その他サイズの基本グループを build_all_mountain_details で統合
        - 各行の工程は入車時間に間に合うように1工程から優先的に割り振り
        - 1工程で間に合わなければ2工程、それでも無理なら3工程、4工程と続ける
        """
        try:
            self.all_mountain_details = build_all_mountain_details(self.group_details, self.size1_mixed_details)
            if self.all_mountain_details is not None and not self.all_mountain_details.empty:
                # まず納入先ベースで仮の工程を付与（明細表示用）
                self.proc_details = compute_proc_details(self.all_mountain_details, self.process_map)
            else:
                self.proc_details = compute_proc_details(self.size1_mixed_details, self.process_map) if (self.size1_mixed_details is not None and not self.size1_mixed_details.empty) else pd.DataFrame()
            
            # 入車時間マスタを読み込み
            master_df = pd.DataFrame()
            try:
                master_path = get_master_path()
                if master_path.exists():
                    master_df = load_pickup_time_master_xlsx(master_path)
            except Exception as e:
                print(f"入車時間マスタ読み込みエラー: {e}")
            
            # 入車時間ベースで工程を自動割り振り
            if not master_df.empty:
                self.mountain_proc = assign_processes_by_arrival_time(self.proc_details, master_df, num_processes=4)
            else:
                # マスタがない場合は従来の優先度ベース
                self.mountain_proc = compute_mountain_process(self.proc_details, strategy="priority", force_takebe_to_3=False)
            
            self.mountain_proc_map = dict(zip(self.mountain_proc["山通番"], self.mountain_proc["山工程"]))
            
            # proc_detailsに新しい工程を反映
            if not self.proc_details.empty and self.mountain_proc_map:
                self.proc_details["工程"] = self.proc_details["山通番"].map(
                    lambda y: str(self.mountain_proc_map.get(int(y), "4"))
                )
            
            self.proc_summary = compute_proc_summary(self.proc_details)
        except Exception as e:
            messagebox.showerror("工程割当エラー", f"工程の再計算に失敗しました: {e}")
    # 工程別かんばんタブ更新（山工程を列に追加し、工程色タグでハイライト）
    def update_kb_views(self):
        # 上段サマリ
        for iid in self.kb_summary.get_children():
            self.kb_summary.delete(iid)
        df = self.proc_summary if self.proc_summary is not None else pd.DataFrame()
        if df is None or df.empty:
            self.kb_detail.configure(columns=())
            for iid in self.kb_detail.get_children():
                self.kb_detail.delete(iid)
            if hasattr(self, "kb_yama_label"):
                self.kb_yama_label.configure(text="選択中の山：- / 山工程：-")
            return
        for _, row in df.iterrows():
            yama = int(row.get("山通番"))
            lab = str(getattr(self, "mountain_proc_map", {}).get(yama, "4"))
            # 表示は GUI 内は従来通り（"1","2","3","4"）で保持
            vals = [
                yama, lab,
                row.get("工程1", 0),
                row.get("工程2", 0),
                row.get("工程3", 0),
                row.get("4工程", 0),
                row.get("合計", 0),
            ]
            tag = ("proc_1" if lab == "1" else "proc_2" if lab == "2" else "proc_3" if lab == "3" else "proc_4",)
            iid = f"kb:{yama}"
            self.kb_summary.insert("", "end", iid=iid, values=vals, tags=tag)
        # 明細クリア＆ラベルリセット
        self.kb_detail.configure(columns=())
        for iid in self.kb_detail.get_children():
            self.kb_detail.delete(iid)
        if hasattr(self, "kb_yama_label"):
            self.kb_yama_label.configure(text="選択中の山：- / 山工程：-")
    # サマリ行選択時の明細表示（山工程ラベル、工程不一致行をオレンジ強調）
    def on_kb_select(self, event=None):
        sel = self.kb_summary.selection()
        if not sel:
            return
        vals = self.kb_summary.item(sel[0], "values")
        if not vals:
            return
        try:
            yama = int(vals[0])  # 1列目が山通番
            y_proc = str(vals[1])  # 2列目が山工程
        except Exception:
            return
        if hasattr(self, "kb_yama_label"):
            self.kb_yama_label.configure(text=f"選択中の山：{yama} / 山工程：{y_proc}")
        det = self.proc_details
        if det is None or det.empty:
            return
        det_y = det.loc[det["山通番"] == yama].copy()
        det_y["山工程"] = y_proc
        cols_pref = [
            "納入先", "山工程", "工程", "工程内No",
            "ストア", "SYUKKASAKI", "NONYUHIBIN", "UKEIRE", "移動工数"
        ]
        cols = [c for c in cols_pref if c in det_y.columns]
        if not cols:
            cols = ["納入先", "山工程", "工程", "工程内No", "移動工数"]
        for iid in self.kb_detail.get_children():
            self.kb_detail.delete(iid)
        self.kb_detail.configure(columns=cols, show="headings")
        for c in cols:
            self.kb_detail.heading(c, text=c)
        for c in cols:
            if c == "納入先":
                w = 200
            elif c == "ストア":
                w = 160
            else:
                w = 120
            self.kb_detail.column(c, width=w, anchor="w")
        for i, (_, row) in enumerate(det_y.sort_values(by=["工程", "工程内No"]).iterrows(), 1):
            r_proc = str(row.get("工程", ""))
            if r_proc == "4":
                tags = ("proc_4",)
            else:
                tags = ("proc_mismatch",) if r_proc != y_proc else ()
            self.kb_detail.insert(
                "",
                "end",
                iid=f"kd:{i}",
                values=[row.get(c, "") for c in cols],
                tags=tags
            )
    # Total工数タブ更新
    def update_total_views(self):
        df = compute_total_work_table(
            self.group_details,
            self.group_results,
            self.height_cap.get(),
            self.size1_mixed_summary,
            self.size1_mixed_details
        )
        for iid in getattr(self, "total_tree", ttk.Treeview()).get_children():
            try:
                self.total_tree.delete(iid)
            except Exception:
                pass
        if df is None or df.empty:
            try:
                self.total_tree.configure(columns=("区分","キー","パレット数","total工数","納入先一覧"), show="headings")
            except Exception:
                pass
            return
        cols = list(self.total_tree["columns"])
        for _, row in df.iterrows():
            values = [row.get(c, "") for c in cols]
            iid = f"tot:{row['区分']}:{row['キー']}"
            self.total_tree.insert("", "end", iid=iid, values=values)
    
    # ===== 入車時間マスタ管理メソッド =====
    def load_master(self):
        """入車時間マスタを読み込んでTreeviewに表示"""
        try:
            master_path = get_master_path()
            self.master_data = load_pickup_time_master_xlsx(master_path)
            self.refresh_master_tree()
            messagebox.showinfo("マスタ読込", f"入車時間マスタを読み込みました。\n件数: {len(self.master_data)}件\nパス: {master_path}")
        except Exception as e:
            messagebox.showerror("マスタ読込エラー", f"入車時間マスタの読み込みに失敗しました:\n{e}")
    
    def save_master(self):
        """現在のTreeviewの内容を入車時間マスタファイルに保存"""
        try:
            # Treeviewから全データを取得
            rows = []
            for iid in self.master_tree.get_children():
                values = self.master_tree.item(iid, "values")
                if len(values) >= 3:
                    rows.append({
                        "OData_納入先": str(values[0]).strip(),
                        "NONYUHIBIN": str(values[1]).strip(),
                        "入車時間": str(values[2]).strip()
                    })
            if not rows:
                messagebox.showwarning("マスタ保存", "保存するデータがありません。")
                return
            self.master_data = pd.DataFrame(rows)
            master_path = get_master_path()
            save_pickup_time_master_xlsx(self.master_data, master_path)
            messagebox.showinfo("マスタ保存", f"入車時間マスタを保存しました。\n件数: {len(self.master_data)}件\nパス: {master_path}")
        except Exception as e:
            messagebox.showerror("マスタ保存エラー", f"入車時間マスタの保存に失敗しました:\n{e}")
    
    def refresh_master_tree(self):
        """master_dataの内容をTreeviewに反映"""
        for iid in self.master_tree.get_children():
            self.master_tree.delete(iid)
        if self.master_data is None or self.master_data.empty:
            return
        for i, row in self.master_data.iterrows():
            values = [
                str(row.get("OData_納入先", "")),
                str(row.get("NONYUHIBIN", "")),
                str(row.get("入車時間", ""))
            ]
            self.master_tree.insert("", "end", iid=f"m:{i}", values=values)
    
    def add_master_row(self):
        """新しい行を追加（ダイアログで入力）"""
        dialog = tk.Toplevel(self)
        dialog.title("入車時間マスタ - 行追加")
        dialog.geometry("400x200")
        dialog.transient(self)
        dialog.grab_set()
        
        ttk.Label(dialog, text="OData_納入先:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        dest_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=dest_var, width=30).grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="NONYUHIBIN (便番号):").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        bin_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=bin_var, width=30).grid(row=1, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="入車時間 (HH:MM):").grid(row=2, column=0, padx=10, pady=10, sticky="w")
        time_var = tk.StringVar()
        ttk.Entry(dialog, textvariable=time_var, width=30).grid(row=2, column=1, padx=10, pady=10)
        
        def do_add():
            dest = dest_var.get().strip()
            bin_num = bin_var.get().strip()
            time_str = time_var.get().strip()
            if not dest or not bin_num or not time_str:
                messagebox.showwarning("入力エラー", "全ての項目を入力してください。")
                return
            # NONYUHIBIN を2桁ゼロ埋め
            try:
                bin_int = int(bin_num)
                bin_num = f"{bin_int:02d}"
            except:
                pass
            # Treeviewに追加
            new_id = f"m:new_{len(self.master_tree.get_children())}"
            self.master_tree.insert("", "end", iid=new_id, values=(dest, bin_num, time_str))
            dialog.destroy()
        
        ttk.Button(dialog, text="追加", command=do_add).grid(row=3, column=0, columnspan=2, pady=20)
    
    def delete_master_row(self):
        """選択した行を削除"""
        selection = self.master_tree.selection()
        if not selection:
            messagebox.showinfo("削除", "削除する行を選択してください。")
            return
        if not messagebox.askyesno("確認", f"{len(selection)}行を削除しますか？"):
            return
        for iid in selection:
            self.master_tree.delete(iid)
    
    def clear_master(self):
        """全てのマスタデータをクリア"""
        if not messagebox.askyesno("確認", "全てのマスタデータをクリアしますか？"):
            return
        for iid in self.master_tree.get_children():
            self.master_tree.delete(iid)
        self.master_data = pd.DataFrame(columns=["OData_納入先", "NONYUHIBIN", "入車時間"])
    
    def edit_master_row(self, event=None):
        """ダブルクリックで行を編集"""
        selection = self.master_tree.selection()
        if not selection:
            return
        iid = selection[0]
        values = self.master_tree.item(iid, "values")
        if len(values) < 3:
            return
        
        dialog = tk.Toplevel(self)
        dialog.title("入車時間マスタ - 行編集")
        dialog.geometry("400x200")
        dialog.transient(self)
        dialog.grab_set()
        
        ttk.Label(dialog, text="OData_納入先:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
        dest_var = tk.StringVar(value=values[0])
        ttk.Entry(dialog, textvariable=dest_var, width=30).grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="NONYUHIBIN (便番号):").grid(row=1, column=0, padx=10, pady=10, sticky="w")
        bin_var = tk.StringVar(value=values[1])
        ttk.Entry(dialog, textvariable=bin_var, width=30).grid(row=1, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="入車時間 (HH:MM):").grid(row=2, column=0, padx=10, pady=10, sticky="w")
        time_var = tk.StringVar(value=values[2])
        ttk.Entry(dialog, textvariable=time_var, width=30).grid(row=2, column=1, padx=10, pady=10)
        
        def do_update():
            dest = dest_var.get().strip()
            bin_num = bin_var.get().strip()
            time_str = time_var.get().strip()
            if not dest or not bin_num or not time_str:
                messagebox.showwarning("入力エラー", "全ての項目を入力してください。")
                return
            # NONYUHIBIN を2桁ゼロ埋め
            try:
                bin_int = int(bin_num)
                bin_num = f"{bin_int:02d}"
            except:
                pass
            self.master_tree.item(iid, values=(dest, bin_num, time_str))
            dialog.destroy()
        
        ttk.Button(dialog, text="更新", command=do_update).grid(row=3, column=0, columnspan=2, pady=20)
# ====== ビルド用の明細 DataFrame 作成関数（Treeview で使う） ======
def build_basic_detail_df(group_details: dict, stype: str, gno: int) -> pd.DataFrame:
    det = group_details.get(str(stype))
    if det is None or det.empty:
        return pd.DataFrame()
    df = det.loc[det["グループ番号"] == int(gno)].copy()
    cols_pref = ["納入先","納入先コード","SYUKKASAKI","NONYUHIBIN","UKEIRE","高さ","移動工数","グループ番号"]
    cols = [c for c in cols_pref if c in df.columns]
    return df[cols].reset_index(drop=True) if cols else df.reset_index(drop=True)
def build_mixed_detail_df(size1_mixed_details: pd.DataFrame, yama: int) -> pd.DataFrame:
    if size1_mixed_details is None or size1_mixed_details.empty:
        return pd.DataFrame()
    df = size1_mixed_details.loc[size1_mixed_details["山通番"] == int(yama)].copy()
    cols_pref = ["納入先","納入先コード","ローカルグループ番号","SYUKKASAKI","NONYUHIBIN","UKEIRE","高さ","移動工数","山通番"]
    cols = [c for c in cols_pref if c in df.columns]
    return df[cols].reset_index(drop=True) if cols else df.reset_index(drop=True)
def build_dest_detail_df(size1_mixed_details: pd.DataFrame, yama: int) -> pd.DataFrame:
    if size1_mixed_details is None or size1_mixed_details.empty:
        return pd.DataFrame()
    df = size1_mixed_details.loc[size1_mixed_details["山通番"] == int(yama)].copy()
    if df.empty:
        return pd.DataFrame()
    out = df.groupby("納入先").agg(
        パレット数=("納入先","count"),
        高さ合計=("高さ","sum"),
        Max移動工数=("移動工数","max")
    ).reset_index().sort_values("納入先")
    return out
# アプリ起動
if __name__ == "__main__":
    app = App()
    app.mainloop()
    # 現状最新1/28