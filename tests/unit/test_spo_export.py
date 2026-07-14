# -*- coding: utf-8 -*-
"""CHかんばんセット — SPO安全出力のユニットテスト"""

import os
import sys
import threading
import time
from pathlib import Path
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from src.services import exporter
from src.services import spo_export


def test_write_via_temp_then_copy_creates_output_and_cleans_temp(tmp_path, monkeypatch):
    df = pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})
    output_path = tmp_path / "out.xlsx"

    forced_tmp_path = tmp_path / "forced_tmp.xlsx"

    def fake_mkstemp(suffix: str):
        fd = os.open(str(forced_tmp_path), os.O_CREAT | os.O_RDWR)
        return fd, str(forced_tmp_path)

    monkeypatch.setattr(spo_export.tempfile, "mkstemp", fake_mkstemp)

    returned = spo_export._write_via_temp_then_copy(df, str(output_path))
    assert returned == str(output_path)
    assert output_path.exists()
    out_df = pd.read_excel(output_path, engine="openpyxl")
    assert list(out_df.columns) == ["A", "B"]
    assert out_df.to_dict(orient="records") == [{"A": 1, "B": "x"}, {"A": 2, "B": "y"}]
    assert not forced_tmp_path.exists()


def test_wait_for_sync_returns_true_for_stable_file(tmp_path):
    p = tmp_path / "stable.xlsx"
    p.write_bytes(b"abc")

    ok = spo_export._wait_for_sync(
        str(p),
        stable_sec=0.0,
        check_interval=0.001,
        timeout=0.2,
    )
    assert ok is True


def test_wait_for_sync_timeout_for_missing_file(monkeypatch):
    monkeypatch.setattr(spo_export.time, "sleep", lambda _: None)

    ticks = {"n": 0}

    def fake_monotonic():
        ticks["n"] += 1
        return ticks["n"] * 0.06

    monkeypatch.setattr(spo_export.time, "monotonic", fake_monotonic)

    ok = spo_export._wait_for_sync(
        "C:/does-not-exist.xlsx",
        stable_sec=0.2,
        check_interval=0.0,
        timeout=0.1,
    )
    assert ok is False


def test_wait_for_sync_timeout_when_file_keeps_changing(monkeypatch):
    monkeypatch.setattr(spo_export.time, "sleep", lambda _: None)

    ticks = {"n": 0}

    def fake_monotonic():
        ticks["n"] += 1
        return ticks["n"] * 0.05

    monkeypatch.setattr(spo_export.time, "monotonic", fake_monotonic)

    stats = {"n": 0}

    def fake_stat(_path: str):
        stats["n"] += 1
        return SimpleNamespace(st_size=100, st_mtime=float(stats["n"]))

    monkeypatch.setattr(spo_export.os, "stat", fake_stat)

    ok = spo_export._wait_for_sync(
        "dummy.xlsx",
        stable_sec=0.2,
        check_interval=0.0,
        timeout=0.15,
    )
    assert ok is False


def test_export_lock_serializes_parallel_calls(tmp_path, monkeypatch):
    df = pd.DataFrame({"x": [1]})
    output_path = tmp_path / "serialized.xlsx"

    active = {"count": 0, "max": 0}
    counters_lock = threading.Lock()

    def fake_write(_df: pd.DataFrame, output_path: str):
        with counters_lock:
            active["count"] += 1
            if active["count"] > active["max"]:
                active["max"] = active["count"]
        # lockの直列化が無いとここが重なる
        time.sleep(0.02)
        Path(output_path).write_text("ok", encoding="utf-8")
        with counters_lock:
            active["count"] -= 1
        return output_path

    monkeypatch.setattr(spo_export, "_write_via_temp_then_copy", fake_write)
    monkeypatch.setattr(spo_export, "_wait_for_sync", lambda *_args, **_kwargs: True)

    threads = [
        threading.Thread(target=spo_export.export_to_spo, args=(df, str(output_path)))
        for _ in range(6)
    ]
    for t in threads:
        t.start()
    for t in threads:
        t.join()

    assert active["max"] == 1
    assert output_path.exists()


def test_export_spo_xlsx_adds_spoexport_table(tmp_path):
    spo_df = pd.DataFrame(
        [
            {
                "タイトル": "山1",
                "工程": "1工程",
                "groupdata": "[]",
            }
        ]
    )
    out_path = exporter.export_spo_xlsx(spo_df, str(tmp_path), base_name="SPO_test")

    wb = load_workbook(out_path)
    ws = wb.active
    assert "SPOExport" in ws.tables


def test_spo_xlsx_goes_to_watch_dir_but_history_and_unmatched_go_local():
    """SPO出力先分離: 監視フォルダにはXlsxのみ、履歴とCSVはローカル固定フォルダ"""
    from src.app.gui import resolve_spo_output_dirs, LOCAL_OUTPUT_DIR
    watch = r"C:\OneDrive\SPO監視"
    dirs = resolve_spo_output_dirs(watch)
    assert dirs["spo_xlsx_dir"] == watch
    assert dirs["history_dir"] == LOCAL_OUTPUT_DIR
    assert dirs["unmatched_dir"] == LOCAL_OUTPUT_DIR
    assert dirs["history_dir"] != watch
    assert dirs["unmatched_dir"] != watch


def test_generate_unique_filename_creates_unique_names():
    """タイムスタンプ+UUID4先頭8桁で一意なファイル名を生成する。"""
    name1 = spo_export._generate_unique_filename("SPOアップロード用")
    name2 = spo_export._generate_unique_filename("SPOアップロード用")
    
    # ファイル名のパターンを確認
    assert name1.startswith("SPOアップロード用_")
    assert name1.endswith(".xlsx")
    assert name2.startswith("SPOアップロード用_")
    assert name2.endswith(".xlsx")
    
    # 異なる名前が生成されることを確認
    assert name1 != name2


def test_export_to_spo_staged_empty_dataframe_returns_none(tmp_path):
    """空DataFrameの場合、ファイルを作成せず None を返す。"""
    df_empty = pd.DataFrame()
    watch_dir = tmp_path / "watch"
    staging_dir = tmp_path / "staging"
    
    result = spo_export.export_to_spo_staged(
        df_empty, 
        watch_dir=str(watch_dir),
        staging_dir=str(staging_dir),
        table_name="SPOExport",
        base_name="SPOアップロード用"
    )
    
    assert result is None
    assert not (watch_dir / "*").exists()
    assert not (staging_dir / "*").exists()


def test_export_to_spo_staged_creates_file_in_watch_dir_after_move(tmp_path):
    """staging で完成させ、os.replace で watch_dir に移動。その後存在確認。"""
    df = pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})
    watch_dir = tmp_path / "watch"
    staging_dir = tmp_path / "staging"
    
    result = spo_export.export_to_spo_staged(
        df,
        watch_dir=str(watch_dir),
        staging_dir=str(staging_dir),
        table_name="SPOExport",
        base_name="SPOアップロード用"
    )
    
    # watch_dir に一意名のファイルが存在することを確認
    assert result is not None
    watch_file = Path(result)
    assert watch_file.exists()
    assert watch_file.parent == watch_dir
    
    # ファイル名にタイムスタンプが含まれることを確認
    assert "SPOアップロード用_" in watch_file.name
    assert watch_file.name.endswith(".xlsx")
    
    # staging_dir に残存ファイルがないことを確認
    staging_files = list(staging_dir.glob("*"))
    assert len(staging_files) == 0
    
    # Excelファイルにテーブルがあることをを確認（SPOExportテーブル）
    wb = load_workbook(str(watch_file))
    ws = wb.active
    table_names = [t.name for t in ws.tables.values()]
    assert "SPOExport" in table_names


def test_export_to_spo_staged_fallback_to_shutil_move_on_cross_drive(tmp_path, monkeypatch):
    """os.replace で OSError が発生した場合、shutil.move にフォールバック。"""
    df = pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})
    watch_dir = tmp_path / "watch"
    staging_dir = tmp_path / "staging"
    
    replace_called = {"count": 0}
    original_replace = os.replace
    
    def fake_replace(src, dst):
        replace_called["count"] += 1
        if replace_called["count"] == 1:
            raise OSError("cross-device link")
        return original_replace(src, dst)
    
    monkeypatch.setattr(spo_export.os, "replace", fake_replace)
    
    result = spo_export.export_to_spo_staged(
        df,
        watch_dir=str(watch_dir),
        staging_dir=str(staging_dir),
        table_name="SPOExport",
        base_name="SPOアップロード用"
    )
    
    # フォールバックが発動したことを確認
    assert replace_called["count"] >= 1
    
    # watch_dir にファイルが最終的に存在することを確認
    assert result is not None
    assert Path(result).exists()


def test_export_to_spo_staged_cleanup_on_exception(tmp_path, monkeypatch):
    """例外発生時、staging_path・final_path 双方を削除してから re-raise。"""
    df = pd.DataFrame({"A": [1, 2], "B": ["x", "y"]})
    watch_dir = tmp_path / "watch"
    staging_dir = tmp_path / "staging"
    
    # staging_dir作成を失敗させる（テーブル追加失敗を模擬）
    def fake_add_table_exact(path, table_name):
        raise ValueError("Simulated table creation error")
    
    monkeypatch.setattr(spo_export, "_add_table_exact", fake_add_table_exact)
    
    try:
        result = spo_export.export_to_spo_staged(
            df,
            watch_dir=str(watch_dir),
            staging_dir=str(staging_dir),
            table_name="SPOExport",
            base_name="SPOアップロード用"
        )
        assert False, "例外が発生するはずだが発生しなかった"
    except ValueError as e:
        assert "Simulated table creation error" in str(e)
        # watch_dir が clean であることを確認
        watch_files = list(watch_dir.glob("*")) if watch_dir.exists() else []
        assert len(watch_files) == 0
