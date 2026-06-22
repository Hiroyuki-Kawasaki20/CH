# -*- coding: utf-8 -*-
"""CHかんばんセット — データ読み込み・前処理サービス"""

import sys
import json
import re
from datetime import datetime, time
from pathlib import Path
from typing import Optional, Dict, List, Tuple

import pandas as pd
import numpy as np
from openpyxl import load_workbook

from ..models.constants import (
    CONFIG_FILENAME, DEFAULT_MIXING_KEY, HAISHA_VENDOR_MAP,
    BASE_ONE_TIME, MIDDLE_WORK, BASE_PER_PAL,
)
from ..utils.normalizer import (
    _normalize_dest_name, _normalize_route_name, _normalize_hhmm,
    _normalize_ukeire, _ZEN2HAN_DIGIT_COLON,
)
from ..utils.csv_utils import read_csv_ja


# ===== 設定ファイル管理 =====
def get_config_path() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent / CONFIG_FILENAME
    else:
        return Path(__file__).resolve().parents[2] / "config" / CONFIG_FILENAME


def load_config() -> dict:
    config_path = get_config_path()
    if config_path.exists():
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_config(config: dict):
    config_path = get_config_path()
    config_path.parent.mkdir(parents=True, exist_ok=True)
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def select_data_folder() -> Optional[Path]:
    from tkinter import filedialog
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    folder = filedialog.askdirectory(
        title="データフォルダを選択してください（出荷情報_全便_最新版.csv と 出荷場一覧.csv があるフォルダ）"
    )
    root.destroy()
    if folder:
        return Path(folder)
    return None


def get_base_dir() -> Path:
    from tkinter import messagebox
    config = load_config()
    base_dir_str = config.get("base_dir")
    if base_dir_str:
        base_dir = Path(base_dir_str)
        if base_dir.exists():
            return base_dir
    messagebox.showinfo("初期設定", "データフォルダを選択してください。\n（出荷情報_CH_最新版.csv または 出荷情報_全便_最新版.csv と 出荷場一覧.csv があるフォルダ）")
    while True:
        base_dir = select_data_folder()
        if base_dir is None:
            if messagebox.askyesno("確認", "フォルダが選択されていません。終了しますか？"):
                raise SystemExit("フォルダが選択されませんでした")
            continue
        s_path_ch = base_dir / "出荷情報_CH_最新版.csv"
        s_path_all = base_dir / "出荷情報_全便_最新版.csv"
        p_path = base_dir / "出荷場一覧.csv"
        if (not s_path_ch.exists() and not s_path_all.exists()) or not p_path.exists():
            missing = []
            if not s_path_ch.exists() and not s_path_all.exists():
                missing.append("出荷情報_CH_最新版.csv / 出荷情報_全便_最新版.csv")
            if not p_path.exists():
                missing.append("出荷場一覧.csv")
            messagebox.showerror("エラー", f"必要なファイルがありません:\n{', '.join(missing)}")
            continue
        config["base_dir"] = str(base_dir)
        save_config(config)
        return base_dir


def select_export_folder() -> Optional[Path]:
    from tkinter import filedialog
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    folder = filedialog.askdirectory(
        title="出力先フォルダを選択してください（SPO用Excelを出力するOneDrive共有フォルダ）"
    )
    root.destroy()
    if folder:
        return Path(folder)
    return None


def get_export_dir() -> Path:
    """出力先フォルダを設定ファイルから取得。無い場合はダイアログで選択＆保存"""
    from tkinter import messagebox
    config = load_config()
    export_dir_str = config.get("export_dir")
    if export_dir_str:
        export_dir = Path(export_dir_str)
        if export_dir.exists():
            return export_dir
    messagebox.showinfo("初期設定", "出力先フォルダを選択してください。\n（SPO用Excelを出力するOneDrive共有フォルダ）")
    while True:
        export_dir = select_export_folder()
        if export_dir is None:
            if messagebox.askyesno("確認", "フォルダが選択されていません。終了しますか？"):
                raise SystemExit("出力先フォルダが選択されませんでした")
            continue
        # 出力先の場合は存在確認のみ（書き込み権限は動的にチェック可）
        if not export_dir.is_dir():
            messagebox.showerror("エラー", "指定されたパスはフォルダではありません。")
            continue
        config["export_dir"] = str(export_dir)
        save_config(config)
        return export_dir


def _resolve_shipments_path(base_dir: Path) -> Path:
    """CH版CSVを優先して出荷情報ファイルを決定する"""
    candidates = [
        base_dir / "出荷情報_CH_最新版.csv",
        base_dir / "出荷情報_全便_最新版.csv",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(
        f"出荷情報CSVが見つかりません: {candidates[0]} または {candidates[1]}"
    )


def _supplement_sample_columns(df_shipments: pd.DataFrame, raw_move_series: Optional[pd.Series] = None) -> pd.DataFrame:
    """欠損列を補完しつつ、移動工数は入力された実値をそのまま使う。"""
    df = df_shipments.copy()

    # 元移動工数: 補完前の値を保持
    raw_move_num = None
    if raw_move_series is not None:
        raw_move_num = pd.to_numeric(raw_move_series, errors="coerce")

    if "元移動工数" not in df.columns:
        if raw_move_num is not None:
            df["元移動工数"] = raw_move_num
        elif "移動工数" in df.columns:
            df["元移動工数"] = pd.to_numeric(df["移動工数"], errors="coerce")
        else:
            df["元移動工数"] = np.nan

    # サイズ種類: 欠損時はテスト用に 1 を補完
    if "サイズ種類" not in df.columns:
        df["サイズ種類"] = "1"
    else:
        s = df["サイズ種類"].astype(str).str.strip()
        invalid = s.eq("") | s.str.lower().isin(["nan", "none"])
        df["サイズ種類"] = s.mask(invalid, "1")

    # 移動工数: サンプル補完は行わず、入力された実値のみを使用
    if "移動工数" not in df.columns:
        df["移動工数"] = np.nan
    move_num = pd.to_numeric(df["移動工数"], errors="coerce")
    if raw_move_num is not None:
        df["移動工数"] = raw_move_num.astype(float)
    else:
        df["移動工数"] = move_num.astype(float)

    # パレット数は1以上に正規化
    if "PLANKANBANSU" not in df.columns:
        df["PLANKANBANSU"] = 1
    pal = pd.to_numeric(df["PLANKANBANSU"], errors="coerce").fillna(1).clip(lower=1).astype(int)
    df["PLANKANBANSU"] = pal

    # トータル工数: 移動工数がある行のみ計算
    total_cost = np.round(
        df["移動工数"] + BASE_ONE_TIME + ((pal - 1) * MIDDLE_WORK) + (pal * BASE_PER_PAL), 0
    )
    df["トータル工数"] = pd.Series(total_cost, index=df.index)

    return df


def load_data():
    """出荷情報・出荷場一覧CSVの読込と前処理"""
    base_dir = get_base_dir()
    s_path = _resolve_shipments_path(base_dir)
    p_path = base_dir / "出荷場一覧.csv"
    if not s_path.exists() or not p_path.exists():
        raise FileNotFoundError(f"CSVが見つかりません:\n{s_path}\n{p_path}")
    df_shipments = read_csv_ja(s_path)
    df_places = read_csv_ja(p_path)
    raw_move_series = df_shipments["移動工数"].copy() if "移動工数" in df_shipments.columns else None
    # 列名前後スペース除去
    df_shipments.columns = df_shipments.columns.str.strip()
    df_places.columns = df_places.columns.str.strip()
    # 納入先コード補完
    if "納入先コード" not in df_shipments.columns and "SYUKKASAKI" in df_shipments.columns:
        df_shipments["納入先コード"] = df_shipments["SYUKKASAKI"].astype(str)
    # 数値前処理
    for num_col in ["移動工数", "高さ", "PLANKANBANSU"]:
        if num_col in df_shipments.columns:
            if num_col == "PLANKANBANSU":
                df_shipments[num_col] = pd.to_numeric(df_shipments[num_col], errors="coerce").fillna(1).astype(int)
            elif num_col == "移動工数":
                df_shipments[num_col] = pd.to_numeric(df_shipments[num_col], errors="coerce")
            else:
                df_shipments[num_col] = pd.to_numeric(df_shipments[num_col], errors="coerce").fillna(0)

    # テスト用サンプル補完（欠損時のみ）
    df_shipments = _supplement_sample_columns(df_shipments, raw_move_series=raw_move_series)
    # 文字列前処理
    for col in ["SSYUKKA", "SYUKKASAKI", "SYUKKAKOKU", "UKEIRE", "NONYUHIBIN", "サイズ種類", "納入先", "納入先コード"]:
        if col in df_shipments.columns:
            df_shipments[col] = df_shipments[col].astype(str).fillna("")
    for col in ["便名", "受入", "仕入先工区", "納入先コード", "納入先工区"]:
        if col in df_places.columns:
            df_places[col] = df_places[col].astype(str).fillna("")
    # 便名の表記ゆれ補正
    if "便名" in df_places.columns:
        df_places["便名"] = df_places["便名"].map(_normalize_route_name)
    # 出荷場一覧の必須列チェック
    required_places_cols = ["便名", "受入", "仕入先工区", "納入先コード", "納入先工区"]
    missing = [c for c in required_places_cols if c not in df_places.columns]
    if missing:
        if "納入先コード" in missing and "SYUKKASAKI" in df_places.columns:
            df_places["納入先コード"] = df_places["SYUKKASAKI"].astype(str)
            missing = [c for c in required_places_cols if c not in df_places.columns]
        if missing:
            raise ValueError(f"出荷場一覧.csv に必要な列が不足しています: {missing}")
    return df_shipments, df_places


# ===== 候補抽出ユーティリティ =====
class DataManager:
    """出荷情報・出荷場一覧を保持し、フィルタリング・候補抽出を行うクラス"""

    def __init__(self, df_shipments: pd.DataFrame, df_places: pd.DataFrame):
        self.df_shipments = df_shipments
        self.df_places = df_places

    def _fallback_vendor_series(self) -> pd.Series:
        src = self.df_shipments.get("納入先", self.df_shipments.get("SYUKKASAKI", "")).astype(str)
        return src.map(_normalize_dest_name)

    def _fallback_mask(self, route_name: str, receipt: Optional[str] = None, order: Optional[str] = None) -> pd.Series:
        route_norm = _normalize_dest_name(str(route_name))
        vendor_norm = self._fallback_vendor_series()
        mask = (vendor_norm == route_norm)
        if receipt is not None and "UKEIRE" in self.df_shipments.columns:
            shp_u = self.df_shipments["UKEIRE"].apply(_normalize_ukeire)
            mask = mask & (shp_u == _normalize_ukeire(str(receipt)))
        if order is not None and "NONYUHIBIN" in self.df_shipments.columns:
            mask = mask & (self.df_shipments["NONYUHIBIN"].astype(str).str.strip() == str(order).strip())
        return mask

    def _mask_for_place_row(self, row: pd.Series) -> pd.Series:
        place_code = row.get("納入先コード", "")
        place_ukeire = _normalize_ukeire(row["受入"])
        shipment_ukeire_normalized = self.df_shipments["UKEIRE"].apply(_normalize_ukeire)
        mask = (
            (self.df_shipments["SSYUKKA"] == row["仕入先工区"]) &
            (self.df_shipments["納入先コード"] == str(place_code)) &
            (self.df_shipments["SYUKKAKOKU"] == row["納入先工区"]) &
            (shipment_ukeire_normalized == place_ukeire)
        )
        return mask

    def get_routes(self) -> list:
        routes = self.df_places["便名"].astype(str).str.strip().unique().tolist()
        # CH運用では日野EH・武部は便名選択対象外
        routes = [r for r in routes if r and r not in {"日野EH", "武部"}]
        return sorted(routes)

    def get_receipts_for_route(self, route_name: str) -> list:
        receipts = self.df_places.loc[self.df_places["便名"] == route_name, "受入"].unique().tolist()
        return sorted(receipts)

    def get_orders_for_route(self, route_name: str) -> list:
        ps = self.df_places[self.df_places["便名"] == route_name]
        if ps.empty:
            m = self._fallback_mask(route_name)
            if m.sum() == 0:
                return []
            return sorted(self.df_shipments.loc[m, "NONYUHIBIN"].astype(str).unique().tolist(), reverse=True)
        mask_total = None
        for _, row in ps.iterrows():
            m = self._mask_for_place_row(row)
            mask_total = m if mask_total is None else (mask_total | m)
        out = sorted(self.df_shipments.loc[mask_total, "NONYUHIBIN"].astype(str).unique().tolist(), reverse=True)
        if out:
            return out
        m = self._fallback_mask(route_name)
        return sorted(self.df_shipments.loc[m, "NONYUHIBIN"].astype(str).unique().tolist(), reverse=True)

    def get_orders_for_route_receipt(self, route_name: str, receipt: str) -> list:
        ps = self.df_places[(self.df_places["便名"] == route_name) & (self.df_places["受入"] == receipt)]
        if ps.empty:
            m = self._fallback_mask(route_name, receipt=receipt)
            if m.sum() == 0:
                return []
            return sorted(self.df_shipments.loc[m, "NONYUHIBIN"].astype(str).unique().tolist(), reverse=True)
        orders = set()
        for _, row in ps.iterrows():
            m = self._mask_for_place_row(row)
            matched_orders = self.df_shipments.loc[m, "NONYUHIBIN"].unique().tolist()
            orders.update(matched_orders)
        out = sorted([str(o) for o in orders], reverse=True)
        if out:
            return out
        m = self._fallback_mask(route_name, receipt=receipt)
        return sorted(self.df_shipments.loc[m, "NONYUHIBIN"].astype(str).unique().tolist(), reverse=True)

    def get_receipts_for_route_order(self, route_name: str, order: str) -> list:
        ps = self.df_places[self.df_places["便名"] == route_name]
        receipts = set()
        for _, row in ps.iterrows():
            m = self._mask_for_place_row(row) & (self.df_shipments["NONYUHIBIN"] == str(order))
            if m.sum() > 0:
                receipts.add(row["受入"])
        out = sorted(receipts)
        if out:
            return out
        m = self._fallback_mask(route_name, order=order)
        if m.sum() == 0:
            return []
        return sorted(self.df_shipments.loc[m, "UKEIRE"].astype(str).unique().tolist())

    def filter_shipments(self, selections: list) -> pd.DataFrame:
        """selections: list of {"便名","受入","オーダー"}"""
        masks = []
        for sel in selections:
            ps = self.df_places[
                (self.df_places["便名"] == sel["便名"]) & (self.df_places["受入"] == sel["受入"])
            ]
            if ps.empty:
                continue
            sub_mask_total = None
            for _, place_row in ps.iterrows():
                sub_mask = (
                    self._mask_for_place_row(place_row) &
                    (self.df_shipments["NONYUHIBIN"] == sel["オーダー"])
                )
                sub_mask_total = sub_mask if sub_mask_total is None else (sub_mask_total | sub_mask)
            if sub_mask_total is not None:
                masks.append(sub_mask_total)
        if not masks:
            fallback_masks = []
            for sel in selections:
                fb = self._fallback_mask(sel["便名"], receipt=sel["受入"], order=sel["オーダー"])
                fallback_masks.append(fb)
            if not fallback_masks:
                return pd.DataFrame()
            final_fb = fallback_masks[0]
            for fm in fallback_masks[1:]:
                final_fb |= fm
            return self.df_shipments.loc[final_fb].copy()
        final_mask = masks[0]
        for m in masks[1:]:
            final_mask |= m
        out = self.df_shipments.loc[final_mask].copy()
        if not out.empty:
            return out

        fallback_masks = []
        for sel in selections:
            fb = self._fallback_mask(sel["便名"], receipt=sel["受入"], order=sel["オーダー"])
            fallback_masks.append(fb)
        if not fallback_masks:
            return out
        final_fb = fallback_masks[0]
        for fm in fallback_masks[1:]:
            final_fb |= fm
        return self.df_shipments.loc[final_fb].copy()


# ===== 入車時間マスタ管理 =====
def get_master_path() -> Path:
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent / "入車時間マスタ.xlsx"
    else:
        return Path(__file__).resolve().parents[2] / "入車時間マスタ.xlsx"


def load_pickup_time_master_xlsx(master_path: Path) -> pd.DataFrame:
    if not master_path.exists():
        return pd.DataFrame(columns=["OData_納入先", "NONYUHIBIN", "入車時間", "セットありフラグ"])
    df = pd.read_excel(master_path, sheet_name=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    required_cols = ["OData_納入先", "NONYUHIBIN", "入車時間"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"入車時間マスタに必要な列がありません: {', '.join(missing)}")
    # 任意列: セットありフラグ（未設定時は空文字で扱う）
    if "セットありフラグ" not in df.columns:
        df["セットありフラグ"] = ""
    out_cols = ["OData_納入先", "NONYUHIBIN", "入車時間", "セットありフラグ"]
    df = df[out_cols].copy()
    df["OData_納入先"] = df["OData_納入先"].astype(str).str.strip()
    nony = df["NONYUHIBIN"].astype(str).str.translate(_ZEN2HAN_DIGIT_COLON)
    nony_num = pd.to_numeric(nony.str.extract(r"(\d+)")[0], errors="coerce")
    df["NONYUHIBIN"] = nony_num.apply(lambda n: f"{int(n):02d}" if pd.notna(n) else "")
    df["入車時間"] = df["入車時間"].apply(_normalize_hhmm)
    df["セットありフラグ"] = df["セットありフラグ"].astype(str).str.strip()
    return df


def save_pickup_time_master_xlsx(df: pd.DataFrame, master_path: Path):
    df_save = df.copy()
    expected_cols = ["OData_納入先", "NONYUHIBIN", "入車時間", "セットありフラグ"]
    for col in expected_cols:
        if col not in df_save.columns:
            df_save[col] = ""
    df_save = df_save[expected_cols]
    df_save.to_excel(master_path, index=False, engine="openpyxl", sheet_name="入車時間マスタ")


def _normalize_excel_time_value(value) -> str:
    """Excel由来の時刻値を HH:MM に正規化する。"""
    if value is None or pd.isna(value):
        return ""

    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"

    if isinstance(value, (int, float)):
        # Excelシリアル時刻（0.0〜1.0）を分解
        if 0 <= float(value) < 1:
            total_minutes = int(round(float(value) * 24 * 60))
            hh = (total_minutes // 60) % 24
            mm = total_minutes % 60
            return f"{hh:02d}:{mm:02d}"

    s = str(value).strip().translate(_ZEN2HAN_DIGIT_COLON)
    if not s:
        return ""

    normalized = _normalize_hhmm(s)
    if normalized:
        return normalized

    # 例: 830 / 0830 形式も補足
    m = re.fullmatch(r"(\d{1,2})(\d{2})", s)
    if m:
        hh, mm = int(m.group(1)), int(m.group(2))
        if 0 <= hh <= 47 and 0 <= mm <= 59:
            return f"{hh:02d}:{mm:02d}"
    return ""


def _resolve_column_name(columns: List[str], candidates: List[str]) -> Optional[str]:
    """候補名（表記ゆれ含む）から実列名を解決する。"""
    normalized = {
        re.sub(r"[^0-9a-zA-Z一-龯ぁ-ゔァ-ヴー]+", "", str(c)).lower(): str(c)
        for c in columns
    }
    for cand in candidates:
        key = re.sub(r"[^0-9a-zA-Z一-龯ぁ-ゔァ-ヴー]+", "", str(cand)).lower()
        if key in normalized:
            return normalized[key]
    return None


def _detect_header_row_index(raw_df: pd.DataFrame, max_scan_rows: int = 20) -> int:
    """先頭の説明行を飛ばし、実ヘッダー行の候補インデックスを返す。"""
    vendor_candidates = [
        "OData_納入先", "納入先", "仕入先", "取引先", "ベンダー", "メーカー", "便名",
    ]
    bin_candidates = [
        "NONYUHIBIN", "納入便", "便番号", "便No", "便No.", "便NO", "便NO.", "便",
    ]
    time_candidates = [
        "入車時間", "入車時刻", "入射時間", "納入時間", "到着時間", "到着時刻", "時刻", "時間",
    ]
    receipt_candidates = ["受入", "受入先", "受入コード", "受入CD", "受入区分"]

    def _norm(text: str) -> str:
        return re.sub(r"[^0-9a-zA-Z一-龯ぁ-ゔァ-ヴー]+", "", str(text)).lower()

    groups = [
        {_norm(x) for x in vendor_candidates},
        {_norm(x) for x in bin_candidates},
        {_norm(x) for x in time_candidates},
        {_norm(x) for x in receipt_candidates},
    ]

    best_idx = 0
    best_score = -1
    scan_len = min(max_scan_rows, len(raw_df.index))
    for i in range(scan_len):
        row_vals = [
            _norm(v) for v in raw_df.iloc[i].tolist()
            if v is not None and str(v).strip() != ""
        ]
        if not row_vals:
            continue
        keys = set(row_vals)
        score = sum(1 for g in groups if keys & g)
        if score > best_score:
            best_score = score
            best_idx = i
        # 必須の3要素（納入先/便名, 便No, 時刻）を満たしたら即採用
        if score >= 3:
            return i

    return best_idx


def _expand_ch_master_vendors(raw_vendor: str, vendor_map: Optional[dict]) -> List[str]:
    """CH入車時間マスタ向けの便名変換・展開ルール。"""
    base = str(raw_vendor).strip()
    if not base:
        return []

    normalized = base.upper().replace("ー", "-").replace("－", "-").replace("―", "-")

    # 現場指定の固定変換
    if normalized.endswith("-TP") or normalized == "TP":
        return ["日野"]
    if normalized.endswith("-KVC") or normalized == "KVC":
        return ["KVC"]
    if normalized.endswith("-RH") or normalized == "RH":
        return ["元町", "高岡"]
    if base == "三栄本社":
        return ["三栄"]
    if base == "織機成形":
        return ["織機"]

    # まず通常の便名マップ（例: TMK->KVC, 三栄SE->三栄）を適用
    mapped = vendor_map.get(base, base) if vendor_map else base
    return [mapped]


def parse_ukeire_ch_excel(
    file_path: Path,
    sheet_name: str = "全受入_納入便データ",
    vendor_map: Optional[dict] = None,
) -> pd.DataFrame:
    """受入データExcelから、受入=CHの入車時間マスタを抽出する。"""
    if vendor_map is None:
        vendor_map = HAISHA_VENDOR_MAP

    try:
        raw_src = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    except Exception as e:
        raise ValueError(f"シート '{sheet_name}' の読み込みに失敗しました: {e}")

    if raw_src is None or raw_src.empty:
        return pd.DataFrame(columns=["OData_納入先", "NONYUHIBIN", "入車時間"])

    header_idx = _detect_header_row_index(raw_src)
    header_vals = [str(v).strip() if v is not None else "" for v in raw_src.iloc[header_idx].tolist()]

    src = raw_src.iloc[header_idx + 1:].copy()
    src.columns = header_vals
    valid_cols = []
    for c in src.columns:
        s = str(c).strip()
        s_low = s.lower()
        if s == "":
            continue
        if s_low in {"nan", "none"}:
            continue
        if s_low.startswith("unnamed"):
            continue
        valid_cols.append(c)
    src = src.loc[:, valid_cols]
    src = src.dropna(how="all")
    if src.empty:
        return pd.DataFrame(columns=["OData_納入先", "NONYUHIBIN", "入車時間"])

    src.columns = [str(c).strip() for c in src.columns]
    cols = list(src.columns)

    receipt_col = _resolve_column_name(cols, [
        "受入", "受入先", "受入コード", "受入CD", "受入区分",
    ])
    vendor_col = _resolve_column_name(cols, [
        "OData_納入先", "納入先", "仕入先", "取引先", "ベンダー", "メーカー", "便名",
    ])
    bin_col = _resolve_column_name(cols, [
        "NONYUHIBIN", "納入便", "便番号", "便No", "便No.", "便NO", "便NO.", "便",
    ])
    time_col = _resolve_column_name(cols, [
        "入車時間", "入車時刻", "入射時間", "納入時間", "到着時間", "到着時刻", "時刻", "時間",
    ])

    missing = []
    if vendor_col is None:
        missing.append("納入先/便名")
    if bin_col is None:
        missing.append("納入便/便No")
    if time_col is None:
        missing.append("入車時間/到着時間")
    if missing:
        raise ValueError(
            "必要列が見つかりません: " + ", ".join(missing) +
            f"\n検出列: {', '.join(cols)}"
        )

    if receipt_col is None:
        work = src.copy()
    else:
        work = src[src[receipt_col].astype(str).str.strip().str.upper() == "CH"].copy()
    if work.empty:
        return pd.DataFrame(columns=["OData_納入先", "NONYUHIBIN", "入車時間"])

    records: List[dict] = []
    for _, row in work.iterrows():
        raw_vendor = str(row.get(vendor_col, "")).strip()
        raw_bin = str(row.get(bin_col, "")).strip().translate(_ZEN2HAN_DIGIT_COLON)
        raw_time = row.get(time_col, "")

        if not raw_vendor or not raw_bin:
            continue

        time_str = _normalize_excel_time_value(raw_time)
        if not time_str:
            continue

        bin_num = pd.to_numeric(pd.Series([raw_bin]).str.extract(r"(\d+)")[0], errors="coerce").iloc[0]
        if pd.isna(bin_num):
            continue

        vendors = _expand_ch_master_vendors(raw_vendor, vendor_map)
        for vendor in vendors:
            records.append({
                "OData_納入先": vendor,
                "NONYUHIBIN": f"{int(bin_num):02d}",
                "入車時間": time_str,
            })

    if not records:
        return pd.DataFrame(columns=["OData_納入先", "NONYUHIBIN", "入車時間"])

    df = pd.DataFrame(records, columns=["OData_納入先", "NONYUHIBIN", "入車時間"])
    df = df.drop_duplicates(subset=["OData_納入先", "NONYUHIBIN"], keep="first")
    df["_sort_bin"] = pd.to_numeric(df["NONYUHIBIN"], errors="coerce").fillna(0)
    df = df.sort_values(["OData_納入先", "_sort_bin"]).drop(columns=["_sort_bin"]).reset_index(drop=True)
    return df


def parse_haisha_excel(file_path: Path, vendor_map: dict = None) -> pd.DataFrame:
    """配車表Excel（Ｎ８ *.xlsm）から入車時間マスタデータを抽出"""
    if vendor_map is None:
        vendor_map = HAISHA_VENDOR_MAP
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    data: Dict[tuple, object] = {}
    max_row = ws.max_row or 100
    max_col = ws.max_column or 80
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=False):
        for cell in row:
            if cell.value is not None:
                data[(cell.row, cell.column)] = cell.value
    wb.close()
    results: List[dict] = []
    no_bin_vendors: List[dict] = []
    for r in range(1, max_row + 1):
        c_val = data.get((r, 3), "")
        e_val = data.get((r, 5), "")
        if not c_val or not e_val:
            continue
        if not re.match(r'^\d{2}N$', str(c_val).strip()):
            continue
        raw_vendor = str(e_val).strip()
        vendor = vendor_map.get(raw_vendor, raw_vendor)
        times_row = r - 1
        g_val = data.get((times_row, 7), None)
        try:
            total_trips = int(g_val) if g_val is not None else 0
        except (ValueError, TypeError):
            total_trips = 0
        pairs: List[Tuple[str, str]] = []
        arrival_times_only: List[str] = []
        for col in range(11, max_col + 1):
            time_val = data.get((times_row, col))
            bin_val = data.get((r, col))
            if time_val is None:
                continue
            time_str = str(time_val).strip()
            normalized = _normalize_hhmm(time_str)
            if not normalized:
                continue
            if bin_val is not None:
                bin_str = str(bin_val).strip()
                try:
                    bin_num = int(bin_str)
                    pairs.append((normalized, f"{bin_num:02d}"))
                except (ValueError, TypeError):
                    arrival_times_only.append(normalized)
            else:
                arrival_times_only.append(normalized)
        if pairs:
            for time_s, bin_s in pairs:
                results.append({"OData_納入先": vendor, "NONYUHIBIN": bin_s, "入車時間": time_s})
        elif arrival_times_only and total_trips > 0:
            shift = "2直" if r > 36 else "1直"
            no_bin_vendors.append({"vendor": vendor, "times": arrival_times_only,
                                    "total_trips": total_trips, "shift": shift})
    by_vendor: Dict[str, list] = {}
    for info in no_bin_vendors:
        by_vendor.setdefault(info["vendor"], []).append(info)
    for vendor, infos in by_vendor.items():
        sorted_infos = sorted(infos, key=lambda x: 0 if x["shift"] == "2直" else 1)
        bin_counter = 1
        for info in sorted_infos:
            times = info["times"]
            trips = info["total_trips"]
            num_times = len(times)
            if num_times == 0:
                continue
            base, remainder = divmod(trips, num_times)
            for i, t in enumerate(times):
                count = base + (1 if i < remainder else 0)
                for _ in range(count):
                    results.append({"OData_納入先": vendor, "NONYUHIBIN": f"{bin_counter:02d}", "入車時間": t})
                    bin_counter += 1
    df = pd.DataFrame(results, columns=["OData_納入先", "NONYUHIBIN", "入車時間"])
    df["_sort_bin"] = pd.to_numeric(df["NONYUHIBIN"], errors="coerce").fillna(0)
    df = df.sort_values(["OData_納入先", "_sort_bin"]).drop(columns=["_sort_bin"]).reset_index(drop=True)
    return df
