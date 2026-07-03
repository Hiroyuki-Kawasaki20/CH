# -*- coding: utf-8 -*-
"""CHかんばんセット — Excel操作ユーティリティ"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def _ensure_columns(df: pd.DataFrame, cols_order: list) -> pd.DataFrame:
    """DataFrame列の存在確認と順序調整"""
    df2 = df.copy()
    for c in cols_order:
        if c not in df2.columns:
            df2[c] = ""
    return df2.reindex(columns=cols_order + [c for c in df2.columns if c not in cols_order])


def _protect_excel_injection(df: pd.DataFrame, text_cols: list) -> pd.DataFrame:
    """Excelで = + - @ 先頭の文字列が数式化されるのを防止"""
    def safe(s):
        if pd.isna(s):
            return s
        s = str(s)
        return "'" + s if s[:1] in ("=", "+", "-", "@") else s
    df2 = df.copy()
    for c in text_cols:
        if c in df2.columns:
            df2[c] = df2[c].map(safe)
    return df2


def index_to_letters(n: int) -> str:
    """1→A, 26→Z, 27→AA... のExcel列名風変換"""
    if n is None or n <= 0:
        return ""
    letters = []
    while n > 0:
        n -= 1
        letters.append(chr(65 + (n % 26)))
        n //= 26
    return "".join(reversed(letters))


def _add_table_exact(file_path, table_name):
    """Excelファイルにテーブル機能を付与（同名テーブルがあればスキップ）"""
    wb = load_workbook(file_path)
    ws = wb.active
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 1 or max_col < 1:
        wb.save(file_path)
        return
    if table_name in ws.tables:
        wb.save(file_path)
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tbl = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tbl.tableStyleInfo = style
    names = list(ws.tables.keys())
    if table_name in names:
        tbl.displayName = f"{table_name}_{len(names) + 1}"
    ws.add_table(tbl)
    wb.save(file_path)
